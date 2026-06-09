#!/usr/bin/env python3
"""
consultar_notas.py - Consulta XMLs de NFSe (Prefeitura SP) e NF-e (SEFAZ)
usando certificado digital .pfx e gera planilha Excel com todas as notas.

Uso:
    pip install cryptography requests lxml openpyxl
    python consultar_notas.py

Saida:
    xmls_consultados/nfse/  <- XMLs individuais (NFSe)
    xmls_consultados/nfe/   <- XMLs individuais (NF-e)
    notas_fiscais_PERIODO.xlsx  <- planilha consolidada
"""

import os
import sys
import gzip
import base64
import tempfile
import textwrap
import datetime
import xml.etree.ElementTree as ET

import requests
import urllib3
from cryptography.hazmat.primitives.serialization import pkcs12, Encoding, PrivateFormat, NoEncryption
from cryptography.hazmat.backends import default_backend

# Certificados ICP-Brasil nao estao no bundle padrao do Python/certifi.
# verify=False e necessario para conectar nos webservices do governo.
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# == Configuracao ================================================================
PFX_PATH    = 'Cmport_123456.pfx'
PFX_SENHA   = b'123456'

CNPJ        = '22761557000188'
DATA_INICIO = '2026-01-01'
DATA_FIM    = '2026-05-31'

OUTPUT_DIR  = 'xmls_consultados'
# Inscricao municipal da empresa na Prefeitura SP (CCM). Preencher antes de rodar.
INSCRICAO_MUNICIPAL = 'PREENCHER'
# ================================================================================

URL_NFSE_SP   = 'https://nfe.prefeitura.sp.gov.br/ws/lotenfe.asmx'
# NFeDistribuicaoDFe fica no Ambiente Nacional (AN) — www1, nao www
URL_NFE_AN    = 'https://www1.nfe.fazenda.gov.br/NFeDistribuicaoDFe/NFeDistribuicaoDFe.asmx'


# == Certificado =================================================================

def extrair_pem_do_pfx(pfx_path: str, senha: bytes):
    with open(pfx_path, 'rb') as f:
        pfx_data = f.read()
    private_key, certificate, _ = pkcs12.load_key_and_certificates(
        pfx_data, senha, backend=default_backend()
    )
    cert_pem = certificate.public_bytes(Encoding.PEM)
    key_pem  = private_key.private_bytes(Encoding.PEM, PrivateFormat.PKCS8, NoEncryption())
    subject  = certificate.subject
    info = {
        'nome':     subject.get_attributes_for_oid(
                        __import__('cryptography').x509.NameOID.COMMON_NAME)[0].value,
        'validade': certificate.not_valid_after_utc.strftime('%d/%m/%Y'),
    }
    return cert_pem, key_pem, info


def criar_sessao(cert_pem: bytes, key_pem: bytes):
    cert_file = tempfile.NamedTemporaryFile(delete=False, suffix='.pem')
    key_file  = tempfile.NamedTemporaryFile(delete=False, suffix='.pem')
    cert_file.write(cert_pem); cert_file.close()
    key_file.write(key_pem);   key_file.close()
    session = requests.Session()
    session.cert   = (cert_file.name, key_file.name)
    session.verify = False  # ICP-Brasil CA nao esta no certifi
    session.headers.update({'Content-Type': 'text/xml; charset=utf-8'})
    session._temp_files = [cert_file.name, key_file.name]
    return session


def limpar_sessao(session):
    for p in getattr(session, '_temp_files', []):
        try: os.unlink(p)
        except OSError: pass


# == Helpers XML =================================================================

def _iter_tag(root, tag):
    """Retorna todos os elementos com determinada tag (sem namespace)."""
    for el in root.iter():
        t = el.tag.split('}')[-1] if '}' in el.tag else el.tag
        if t == tag:
            yield el


def _first(root, tag):
    return next(_iter_tag(root, tag), None)


def _txt(root, tag, default=''):
    el = _first(root, tag)
    return (el.text or '').strip() if el is not None else default


# == NFSe SP =====================================================================

def consultar_nfse_sp(session, data_inicio: str, data_fim: str):
    """Consulta NFSe emitidas no periodo via Prefeitura SP. Retorna lista de XML str."""
    if INSCRICAO_MUNICIPAL == 'PREENCHER':
        print('\n[NFSe SP] AVISO: preencha INSCRICAO_MUNICIPAL no script e rode novamente.')
        return []

    pedido = textwrap.dedent(f"""\
        <?xml version="1.0" encoding="UTF-8"?>
        <PedidoConsultaNFePeriodo
            xmlns="http://www.prefeitura.sp.gov.br/nfe"
            xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
            xsi:schemaLocation="http://www.prefeitura.sp.gov.br/nfe
                                 http://www.prefeitura.sp.gov.br/nfe/PedidoConsultaNFePeriodo_v01.xsd"
            Versao="1">
          <Cabecalho Versao="1">
            <CPFCNPJRemetente><CNPJ>{CNPJ}</CNPJ></CPFCNPJRemetente>
            <InscricaoMunicipalPrestador>{INSCRICAO_MUNICIPAL}</InscricaoMunicipalPrestador>
            <dtInicio>{data_inicio}</dtInicio>
            <dtFim>{data_fim}</dtFim>
            <NumeroPagina>1</NumeroPagina>
          </Cabecalho>
        </PedidoConsultaNFePeriodo>""")

    envelope = textwrap.dedent(f"""\
        <?xml version="1.0" encoding="utf-8"?>
        <soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/"
                       xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                       xmlns:xsd="http://www.w3.org/2001/XMLSchema">
          <soap:Body>
            <ConsultaNFePeriodo xmlns="http://www.prefeitura.sp.gov.br/nfe">
              <VersaoSchema>1</VersaoSchema>
              <MensagemXML><![CDATA[{pedido}]]></MensagemXML>
            </ConsultaNFePeriodo>
          </soap:Body>
        </soap:Envelope>""")

    print(f'\n[NFSe SP] Consultando {data_inicio} a {data_fim} ...')
    try:
        resp = session.post(
            URL_NFSE_SP,
            data=envelope.encode('utf-8'),
            headers={'SOAPAction': '"http://www.prefeitura.sp.gov.br/nfe/ws/consultaNFePeriodo"'},
            timeout=30,
        )
        print(f'         Status HTTP: {resp.status_code}')
        if resp.status_code != 200:
            print(f'         Erro: {resp.text[:500]}')
            return []

        _salvar(OUTPUT_DIR, 'nfse', 'resposta_bruta.xml', resp.text)
        xmls = _extrair_comp_nfse(resp.text)
        print(f'         {len(xmls)} nota(s) encontrada(s).')
        return xmls
    except requests.exceptions.SSLError as e:
        print(f'         Erro SSL: {e}')
        return []
    except Exception as e:
        print(f'         Erro: {e}')
        return []


def _extrair_comp_nfse(xml_resposta: str):
    xmls = []
    try:
        root = ET.fromstring(xml_resposta)
        for el in _iter_tag(root, 'CompNfse'):
            xmls.append(ET.tostring(el, encoding='unicode'))
    except ET.ParseError:
        import re
        xmls = re.findall(r'<CompNfse>.*?</CompNfse>', xml_resposta, re.DOTALL)
    return xmls


def parsear_nfse(xml_str: str) -> dict:
    """Extrai campos de uma CompNfse para linha da planilha."""
    try:
        root = ET.fromstring(xml_str)

        # Tomador - busca dentro de TomadorServico
        tom_cnpj = tom_nome = ''
        tom_el = _first(root, 'TomadorServico')
        if tom_el is not None:
            cnpj_el = _first(tom_el, 'Cnpj')
            cpf_el  = _first(tom_el, 'Cpf')
            tom_cnpj = (cnpj_el.text or '').strip() if cnpj_el is not None else \
                       (cpf_el.text  or '').strip() if cpf_el  is not None else ''
            rs_el = _first(tom_el, 'RazaoSocial') or _first(tom_el, 'NomeFantasia')
            tom_nome = (rs_el.text or '').strip() if rs_el is not None else ''

        data_raw = _txt(root, 'DataEmissao')
        valor_raw = _txt(root, 'ValorServicos')

        try:
            valor = float(valor_raw) if valor_raw else 0.0
        except ValueError:
            valor = 0.0

        return {
            'Numero':         _txt(root, 'Numero'),
            'Data Emissao':   data_raw[:10] if data_raw else '',
            'Tomador CNPJ':   tom_cnpj,
            'Tomador Nome':   tom_nome,
            'Valor (R$)':     valor,
            'Discriminacao':  _txt(root, 'Discriminacao')[:400],
            'Cod Verificacao': _txt(root, 'CodigoVerificacao'),
        }
    except Exception as e:
        return {'Numero': f'ERRO_PARSE: {e}', 'Data Emissao': '', 'Tomador CNPJ': '',
                'Tomador Nome': '', 'Valor (R$)': 0.0, 'Discriminacao': '', 'Cod Verificacao': ''}


# == NF-e SEFAZ ==================================================================

def consultar_nfe_sefaz(session, ult_nsu: str = '000000000000000'):
    """Consulta DF-e paginando ate o fim. Retorna lista de dicts {nsu, schema, xml}."""
    todos = []
    pagina = 1
    soap_action = '"http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe/nfeDistDFeInteresse"'
    urls = [URL_NFE_AN]

    while True:
        print(f'\n[NF-e SEFAZ] Pagina {pagina} - ultNSU: {ult_nsu} ...')
        envelope = _envelope_dfe(ult_nsu)
        resp = None

        for url in urls:
            try:
                r = session.post(
                    url,
                    data=envelope.encode('utf-8'),
                    headers={'Content-Type': 'text/xml; charset=utf-8',
                             'SOAPAction': soap_action},
                    timeout=60,
                )
                print(f'             {url.split("/")[2]} -> HTTP {r.status_code}')
                if r.status_code == 200:
                    resp = r
                    break
                else:
                    print(f'             Resposta: {r.text[:200]}')
            except Exception as e:
                print(f'             {url.split("/")[2]} -> Erro: {e}')

        if resp is None:
            print('             Nenhum endpoint respondeu. Encerrando.')
            break

        _salvar(OUTPUT_DIR, 'nfe', f'resposta_pag{pagina:03d}.xml', resp.text)
        docs, proximo_nsu, tem_mais = _extrair_docs_dfe(resp.text)
        todos.extend(docs)
        print(f'             {len(docs)} doc(s) nesta pagina. Total: {len(todos)}')

        if not tem_mais or proximo_nsu == ult_nsu:
            break
        ult_nsu = proximo_nsu
        pagina += 1

    return todos


def _envelope_dfe(ult_nsu: str) -> str:
    return textwrap.dedent(f"""\
        <?xml version="1.0" encoding="utf-8"?>
        <soap:Envelope xmlns:soap="http://schemas.xmlsoap.org/soap/envelope/"
                       xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance"
                       xmlns:xsd="http://www.w3.org/2001/XMLSchema">
          <soap:Body>
            <nfeDistDFeInteresse xmlns="http://www.portalfiscal.inf.br/nfe/wsdl/NFeDistribuicaoDFe">
              <nfeDadosMsg>
                <distDFeInt xmlns="http://www.portalfiscal.inf.br/nfe" versao="1.01">
                  <tpAmb>1</tpAmb>
                  <cUFAutor>35</cUFAutor>
                  <CNPJ>{CNPJ}</CNPJ>
                  <distNSU><ultNSU>{ult_nsu}</ultNSU></distNSU>
                </distDFeInt>
              </nfeDadosMsg>
            </nfeDistDFeInteresse>
          </soap:Body>
        </soap:Envelope>""")


def _extrair_docs_dfe(xml_resposta: str):
    """Extrai documentos comprimidos do DF-e. Retorna (docs, proximo_nsu, tem_mais)."""
    docs = []
    proximo_nsu = '000000000000000'
    ult_retornado = '000000000000000'

    try:
        root = ET.fromstring(xml_resposta)
    except ET.ParseError as e:
        print(f'             Parse error na resposta: {e}')
        return docs, proximo_nsu, False

    max_nsu = '000000000000000'
    for el in _iter_tag(root, 'maxNSU'):
        if el.text:
            max_nsu = el.text.strip().zfill(15)
    for el in _iter_tag(root, 'ultNSU'):
        if el.text:
            ult_retornado = el.text.strip().zfill(15)

    # proximo_nsu = ultNSU da resposta (parâmetro para a próxima requisição)
    # tem_mais    = ainda há documentos entre ultNSU e maxNSU
    proximo_nsu = ult_retornado
    tem_mais = max_nsu > ult_retornado

    for el in _iter_tag(root, 'docZip'):
        if not el.text:
            continue
        nsu    = el.attrib.get('NSU', '0').zfill(15)
        schema = el.attrib.get('schema', '')
        try:
            xml_bytes = gzip.decompress(base64.b64decode(el.text.strip()))
            xml_str   = xml_bytes.decode('utf-8')
            docs.append({'nsu': nsu, 'schema': schema, 'xml': xml_str})
        except Exception as e:
            print(f'             Aviso: nao descompactou NSU {nsu}: {e}')

    return docs, proximo_nsu, tem_mais


_EVENTO_MAP = {
    '110111': 'Cancelamento',
    '110110': 'Carta de Correcao',
    '110113': 'EPEC',
    '210200': 'Confirmacao de Operacao',
    '210210': 'Ciencia de Operacao',
    '210220': 'Desconhecimento de Operacao',
    '210240': 'Operacao nao Realizada',
}

_SIT_MAP = {'1': 'Autorizada', '2': 'Cancelada', '3': 'Denegada'}


def parsear_nfe(doc: dict) -> dict:
    """Extrai campos de um doc DF-e (procNFe, resNFe ou procEventoNFe) para linha da planilha."""
    nsu    = doc['nsu']
    schema = doc['schema']
    xml_str = doc['xml']

    base = {
        'NSU': nsu, 'Tipo Doc': schema, 'Chave': '', 'Serie': '', 'Numero': '',
        'Data Emissao': '', 'Tipo NF': '', 'Emitente CNPJ': '', 'Emitente Nome': '',
        'Dest CNPJ': '', 'Dest Nome': '', 'Valor (R$)': 0.0, 'Status': '',
    }

    try:
        root = ET.fromstring(xml_str)
    except ET.ParseError:
        base['Tipo Doc'] = f'ERRO_PARSE:{schema}'
        return base

    if 'resNFe' in schema:
        chave    = _txt(root, 'chNFe')
        data_raw = _txt(root, 'dhEmi')
        tp_nf    = _txt(root, 'tpNF')
        valor_raw = _txt(root, 'vNF')
        try: valor = float(valor_raw)
        except: valor = 0.0
        base.update({
            'Tipo Doc':       'Resumo NF-e',
            'Chave':          chave,
            'Serie':          chave[22:25] if len(chave) >= 44 else '',
            'Numero':         str(int(chave[25:34])) if len(chave) >= 44 else '',
            'Data Emissao':   data_raw[:10] if data_raw else '',
            'Tipo NF':        'Entrada' if tp_nf == '0' else 'Saida',
            'Emitente CNPJ':  _txt(root, 'CNPJ'),
            'Emitente Nome':  _txt(root, 'xNome'),
            'Valor (R$)':     valor,
            'Status':         _SIT_MAP.get(_txt(root, 'cSitNFe'), _txt(root, 'cSitNFe')),
        })
        return base

    if 'procEventoNFe' in schema or 'evento' in schema.lower():
        tp_ev = _txt(root, 'tpEvento')
        dh_ev = _txt(root, 'dhEvento') or _txt(root, 'dhRegEvento')
        base.update({
            'Tipo Doc':     f'Evento: {_EVENTO_MAP.get(tp_ev, tp_ev or schema)}',
            'Chave':        _txt(root, 'chNFe'),
            'Data Emissao': dh_ev[:10] if dh_ev else '',
            'Status':       'Evento',
        })
        return base

    # procNFe (NF-e completa com protocolo)
    infnfe = _first(root, 'infNFe')
    if infnfe is None:
        base['Tipo Doc'] = f'NF-e (sem infNFe):{schema}'
        base['Chave'] = _txt(root, 'chNFe')
        return base

    chave_el = _first(infnfe, 'chNFe')
    if chave_el is None:
        prot_el  = _first(root, 'protNFe')
        chave_el = _first(prot_el if prot_el is not None else root, 'chNFe')
    chave = (chave_el.text or '').strip() if chave_el is not None else ''

    n_nf  = _txt(infnfe, 'nNF')
    serie = _txt(infnfe, 'serie')
    dh    = _txt(infnfe, 'dhEmi')
    tp_nf = _txt(infnfe, 'tpNF')

    emit_el = _first(infnfe, 'emit')
    emit_cnpj = emit_nome = ''
    if emit_el is not None:
        emit_cnpj = _txt(emit_el, 'CNPJ')
        emit_nome = _txt(emit_el, 'xNome')

    dest_el = _first(infnfe, 'dest')
    dest_cnpj = dest_nome = ''
    if dest_el is not None:
        dest_cnpj = _txt(dest_el, 'CNPJ') or _txt(dest_el, 'CPF')
        dest_nome = _txt(dest_el, 'xNome')

    valor_raw = _txt(infnfe, 'vNF')
    try: valor = float(valor_raw)
    except: valor = 0.0

    # Status via protocolo
    prot_el = _first(root, 'infProt')
    c_stat = _txt(prot_el, 'cStat') if prot_el is not None else ''
    if c_stat == '100':
        status = 'Autorizada'
    elif c_stat in ('101', '151', '155'):
        status = 'Cancelada'
    else:
        status = c_stat or 'Autorizada'

    base.update({
        'Tipo Doc':       'NF-e',
        'Chave':          chave,
        'Serie':          serie,
        'Numero':         str(int(n_nf)) if n_nf.isdigit() else n_nf,
        'Data Emissao':   dh[:10] if dh else '',
        'Tipo NF':        'Entrada' if tp_nf == '0' else 'Saida',
        'Emitente CNPJ':  emit_cnpj,
        'Emitente Nome':  emit_nome,
        'Dest CNPJ':      dest_cnpj,
        'Dest Nome':      dest_nome,
        'Valor (R$)':     valor,
        'Status':         status,
    })
    return base


# == Excel =======================================================================

def gerar_excel(nfse_rows: list, nfe_rows: list, caminho: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()

    azul_escuro = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    fonte_branca = Font(bold=True, color='FFFFFF', size=11)
    fonte_link   = Font(color='1F4E79', underline='single')
    centro       = Alignment(horizontal='center', vertical='center', wrap_text=False)
    esquerda     = Alignment(horizontal='left',   vertical='center', wrap_text=False)
    borda_fina   = Border(
        bottom=Side(style='thin', color='DDDDDD'),
        right =Side(style='thin', color='DDDDDD'),
    )

    def estilizar_header(ws, cols):
        for c, nome in enumerate(cols, 1):
            cell = ws.cell(1, c, nome)
            cell.font      = fonte_branca
            cell.fill      = azul_escuro
            cell.alignment = centro
        ws.row_dimensions[1].height = 20

    def autowidth(ws, cols, max_w=70):
        for c, nome in enumerate(cols, 1):
            col_letter = ws.cell(1, c).column_letter
            best = len(nome) + 2
            for row in ws.iter_rows(min_row=2, min_col=c, max_col=c):
                v = str(row[0].value or '')
                best = max(best, len(v) + 2)
            ws.column_dimensions[col_letter].width = min(best, max_w)

    def zebra(ws, rows_count, ncols):
        cinza = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
        for r in range(2, rows_count + 2):
            fill = cinza if r % 2 == 0 else None
            for c in range(1, ncols + 1):
                cell = ws.cell(r, c)
                if fill:
                    cell.fill = fill
                cell.border = borda_fina

    # ---- Aba NFSe ----
    ws_nfse = wb.active
    ws_nfse.title = 'NFSe'
    cols_nfse = ['Numero', 'Data Emissao', 'Tomador CNPJ', 'Tomador Nome',
                 'Valor (R$)', 'Discriminacao', 'Cod Verificacao']
    estilizar_header(ws_nfse, cols_nfse)
    ws_nfse.freeze_panes = 'A2'

    for ri, row in enumerate(nfse_rows, 2):
        for ci, key in enumerate(cols_nfse, 1):
            v = row.get(key, '')
            cell = ws_nfse.cell(ri, ci, v)
            cell.alignment = esquerda
            if key == 'Valor (R$)' and isinstance(v, float):
                cell.number_format = '#,##0.00'
            elif key == 'Data Emissao':
                cell.alignment = centro

    zebra(ws_nfse, len(nfse_rows), len(cols_nfse))
    autowidth(ws_nfse, cols_nfse)

    # ---- Aba NF-e ----
    ws_nfe = wb.create_sheet('NF-e')
    cols_nfe = ['NSU', 'Tipo Doc', 'Numero', 'Serie', 'Data Emissao', 'Tipo NF',
                'Emitente CNPJ', 'Emitente Nome', 'Dest CNPJ', 'Dest Nome',
                'Valor (R$)', 'Status', 'Chave']
    estilizar_header(ws_nfe, cols_nfe)
    ws_nfe.freeze_panes = 'A2'

    for ri, row in enumerate(nfe_rows, 2):
        for ci, key in enumerate(cols_nfe, 1):
            v = row.get(key, '')
            cell = ws_nfe.cell(ri, ci, v)
            cell.alignment = esquerda
            if key == 'Valor (R$)' and isinstance(v, float):
                cell.number_format = '#,##0.00'
            elif key in ('Data Emissao', 'Tipo NF', 'Status', 'NSU'):
                cell.alignment = centro

    zebra(ws_nfe, len(nfe_rows), len(cols_nfe))
    autowidth(ws_nfe, cols_nfe)

    # ---- Aba Resumo ----
    ws_res = wb.create_sheet('Resumo')
    ws_res.column_dimensions['A'].width = 30
    ws_res.column_dimensions['B'].width = 20

    def add_res(r, label, valor):
        ws_res.cell(r, 1, label).font = Font(bold=True)
        ws_res.cell(r, 2, valor)

    total_nfse_val = sum(r.get('Valor (R$)', 0) or 0 for r in nfse_rows)
    total_nfe_val  = sum(r.get('Valor (R$)', 0) or 0 for r in nfe_rows
                        if r.get('Tipo Doc') in ('NF-e', 'Resumo NF-e'))

    add_res(1,  'Periodo consultado',         f'{DATA_INICIO} a {DATA_FIM}')
    add_res(2,  'CNPJ',                        CNPJ)
    add_res(3,  '',                            '')
    add_res(4,  'NFSe encontradas',            len(nfse_rows))
    add_res(5,  'NFSe - Valor total (R$)',     total_nfse_val)
    ws_res.cell(5, 2).number_format = '#,##0.00'
    add_res(6,  '',                            '')
    add_res(7,  'NF-e documentos totais',      len(nfe_rows))

    nfe_apenas = [r for r in nfe_rows if r.get('Tipo Doc') in ('NF-e', 'Resumo NF-e')]
    nfe_eventos = [r for r in nfe_rows if 'Evento' in (r.get('Tipo Doc') or '')]
    add_res(8,  'NF-e (notas)',                len(nfe_apenas))
    add_res(9,  'NF-e (eventos)',              len(nfe_eventos))
    add_res(10, 'NF-e - Valor total (R$)',     total_nfe_val)
    ws_res.cell(10, 2).number_format = '#,##0.00'
    add_res(11, '',                            '')
    add_res(12, 'Gerado em',                   datetime.datetime.now().strftime('%d/%m/%Y %H:%M'))

    wb.save(caminho)
    print(f'\nPlanilha salva: {caminho}')


# == Utilitarios =================================================================

def _salvar(base_dir: str, subdir: str, nome: str, conteudo: str):
    pasta = os.path.join(base_dir, subdir)
    os.makedirs(pasta, exist_ok=True)
    caminho = os.path.join(pasta, nome)
    with open(caminho, 'w', encoding='utf-8') as f:
        f.write(conteudo)
    return caminho


def salvar_nfse_xmls(xmls: list):
    for i, xml_str in enumerate(xmls, 1):
        nome = f'nfse_{i:04d}.xml'
        _salvar(OUTPUT_DIR, 'nfse', nome, xml_str)
        print(f'    Salvo: {OUTPUT_DIR}/nfse/{nome}')


def salvar_nfe_xmls(docs: list):
    for doc in docs:
        nsu  = doc['nsu']
        tipo = (doc['schema'].split('_')[0]).replace('proc', '').replace('res', 'res_')
        nome = f'nfe_{nsu}_{tipo}.xml'
        _salvar(OUTPUT_DIR, 'nfe', nome, doc['xml'])


# == Carregar XMLs salvos ========================================================

def _carregar_nfe_salvos():
    """Le os XMLs individuais ja salvos em xmls_consultados/nfe/ e retorna lista de docs."""
    pasta = os.path.join(OUTPUT_DIR, 'nfe')
    if not os.path.exists(pasta):
        return []

    docs = []
    for nome in sorted(os.listdir(pasta)):
        if not nome.endswith('.xml') or nome.startswith('resposta_'):
            continue
        caminho = os.path.join(pasta, nome)
        try:
            with open(caminho, encoding='utf-8') as f:
                xml_str = f.read()
            # Extrai NSU do nome do arquivo (nfe_000000000004718_xxx.xml)
            partes = nome.replace('.xml', '').split('_')
            nsu = partes[1] if len(partes) >= 2 else '0'
            # Detecta schema pelo conteudo
            if '<resNFe' in xml_str:
                schema = 'resNFe_v1.01.xsd'
            elif '<procEventoNFe' in xml_str or '<evento' in xml_str.lower():
                schema = 'procEventoNFe_v1.00.xsd'
            else:
                schema = 'procNFe_v4.00.xsd'
            docs.append({'nsu': nsu, 'schema': schema, 'xml': xml_str})
        except Exception as e:
            print(f'         Aviso: nao leu {nome}: {e}')

    return docs


# == Main ========================================================================

def main():
    print('=' * 60)
    print(' CMPort - Consulta de Notas Fiscais por Certificado Digital')
    print('=' * 60)

    if not os.path.exists(PFX_PATH):
        print(f'\nERRO: arquivo {PFX_PATH!r} nao encontrado.')
        sys.exit(1)

    print(f'\n[Cert] Carregando {PFX_PATH} ...')
    try:
        cert_pem, key_pem, info = extrair_pem_do_pfx(PFX_PATH, PFX_SENHA)
        print(f'       Nome    : {info["nome"]}')
        print(f'       Validade: {info["validade"]}')
    except Exception as e:
        print(f'\nERRO ao carregar certificado: {e}')
        sys.exit(1)

    session = criar_sessao(cert_pem, key_pem)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    nfse_rows = []
    nfe_rows  = []

    try:
        # -- NFSe Prefeitura SP --
        print('\n' + '-' * 60)
        print(' PARTE 1 - NFSe (Nota Fiscal de Servico) - Prefeitura SP')
        print('-' * 60)
        xmls_nfse = consultar_nfse_sp(session, DATA_INICIO, DATA_FIM)
        if xmls_nfse:
            salvar_nfse_xmls(xmls_nfse)
            for xml_str in xmls_nfse:
                nfse_rows.append(parsear_nfse(xml_str))

        # -- NF-e SEFAZ --
        print('\n' + '-' * 60)
        print(' PARTE 2 - NF-e (Nota Fiscal Eletronica) - SEFAZ')
        print('-' * 60)
        docs_nfe = consultar_nfe_sefaz(session)
        if docs_nfe:
            salvar_nfe_xmls(docs_nfe)
            for doc in docs_nfe:
                nfe_rows.append(parsear_nfe(doc))
        else:
            # Rate limit ou sem novos docs — usa XMLs ja salvos em disco
            docs_salvos = _carregar_nfe_salvos()
            if docs_salvos:
                print(f'         Usando {len(docs_salvos)} XMLs ja salvos em {OUTPUT_DIR}/nfe/')
                for doc in docs_salvos:
                    nfe_rows.append(parsear_nfe(doc))

    finally:
        limpar_sessao(session)

    # -- Gerar Excel --
    ts = datetime.datetime.now().strftime('%Y%m%d_%H%M')
    nome_xlsx = f'notas_fiscais_{DATA_INICIO[:7]}_a_{DATA_FIM[:7]}_{ts}.xlsx'
    gerar_excel(nfse_rows, nfe_rows, nome_xlsx)

    # -- Resumo --
    print('\n' + '=' * 60)
    print(' Concluido.')
    print(f' NFSe: {len(nfse_rows)} nota(s)')
    print(f' NF-e: {len(nfe_rows)} documento(s)')
    print(f' XMLs : {OUTPUT_DIR}/')
    print(f' Excel: {nome_xlsx}')
    print('=' * 60)


if __name__ == '__main__':
    main()
