"""Export do Fluxo Financeiro completo em .xlsx — pro cliente.

Espelha a LÓGICA da planilha da cliente (regime de caixa: tudo por data de
pagamento), com formato próprio mais limpo. Abas:
  Resumo · Entradas · Saídas · Transferências · Categoria x Mês · (Pendências)

Fonte: FinDashboardService.por_banco (totais), FluxoFinanceiroService.fluxo_mensal
(entradas linha a linha), fin_movimentacoes (saídas/transferências linha a linha).
"""
import calendar
import io
from datetime import date
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from sqlalchemy import func
from sqlalchemy.orm import Session, aliased

from app.models.fin_movimentacao_model import MovimentacaoFinanceira
from app.models.fin_categoria_model import CategoriaFinanceira, GrupoCategoria
from app.models.condominio_model import Condominio
from app.models.banco_model import Banco
from app.models.despesa_model import Despesa, DespesaParcela, StatusParcelaDespesa
from app.models.funcionario_model import Funcionario
from app.services.fin_dashboard_service import FinDashboardService
from app.services.fluxo_financeiro_service import FluxoFinanceiroService

EMPRESA_POR_CNPJ = {"22761557000188": "CMPORT", "65756913000188": "TEC"}
_HDR_FILL = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True)
_TOT_FONT = Font(bold=True)
_TERMOS_TARIFA = ("tarifa", "juros", "iof", "ir sobre", "imposto sobre")


def _sd(v: Optional[str]) -> str:
    return "".join(filter(str.isdigit, v or ""))


def _f(v) -> float:
    return round(float(v or 0), 2)


def _meses(ano_ini, mes_ini, ano_fim, mes_fim):
    a, m = ano_ini, mes_ini
    while (a, m) <= (ano_fim, mes_fim):
        yield a, m
        m += 1
        if m > 12:
            m, a = 1, a + 1


def _cab(ws, headers):
    ws.append(headers)
    for i, _ in enumerate(headers, 1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font = _HDR_FILL, _HDR_FONT
        c.alignment = Alignment(horizontal="center")


def _ajustar_largura(ws, com_filtro=True):
    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max((len(str(c.value or "")) for c in col), default=8) + 2, 55
        )
    ws.freeze_panes = "A2"
    if com_filtro and ws.max_row > 1:
        ws.auto_filter.ref = ws.dimensions


class FinExportService:

    @staticmethod
    def gerar_xlsx(db: Session, ano_ini: int, mes_ini: int, ano_fim: int, mes_fim: int,
                   cnpj: Optional[str] = None, incluir_pendentes: bool = True) -> bytes:
        cnpj_limpo = _sd(cnpj) if cnpj else None
        meses = list(_meses(ano_ini, mes_ini, ano_fim, mes_fim))
        rotulo_mes = {(a, m): f"{m:02d}/{a}" for a, m in meses}

        wb = Workbook()
        wb.remove(wb.active)

        # ── coletas ─────────────────────────────────────────────────────────
        resumo_rows = []
        entradas_rows = []
        saidas_rows = []
        transf_rows = []
        cat_mes: dict[str, dict] = {}   # categoria -> {(a,m): total}

        for a, m in meses:
            rot = rotulo_mes[(a, m)]
            dash = FinDashboardService.por_banco(db, a, m)
            # saldo inicial / extrato por CNPJ (do dashboard por-banco)
            base: dict[str, dict] = {}
            for l in dash.bancos:
                if l.banco_id is None:
                    continue
                emp = l.empresa or "?"
                d = base.setdefault(emp, {"saldo_inicial": 0.0, "saldo_extrato": 0.0,
                                          "tem_extrato": False, "rendimento": 0.0})
                d["saldo_inicial"] += _f(l.saldo_inicial)
                d["rendimento"] += _f(l.rendimento)
                if l.saldo_extrato is not None:
                    d["saldo_extrato"] += _f(l.saldo_extrato)
                    d["tem_extrato"] = True

            # entradas por CNPJ (do fluxo_mensal — totais já quebrados por tipo)
            ent_por_emp: dict[str, dict] = {}
            fx = FluxoFinanceiroService.fluxo_mensal(db, a, m, cnpj)
            for c in fx.cnpjs:
                emp = EMPRESA_POR_CNPJ.get(_sd(c.cnpj), c.cnpj)
                e = ent_por_emp.setdefault(emp, {"boleto": 0.0, "recibo": 0.0})
                e["boleto"] += _f(c.total_manutencao) + _f(c.total_assistencia) + _f(c.total_produto)
                e["recibo"] += _f(c.total_recibos)
                for ln in c.linhas:
                    entradas_rows.append([
                        rot, emp, ln.condominio_nome, ln.tipo, ln.numero_nota or "", ln.origem,
                        ln.data_pagamento.strftime("%d/%m/%Y") if ln.data_pagamento else "",
                        ln.valor, ln.banco_nome or "", "cross-CNPJ" if ln.cross_cnpj else "",
                    ])
                    chave = f"Entrada — {ln.tipo.capitalize()}"
                    cat_mes.setdefault(chave, {})
                    cat_mes[chave][(a, m)] = round(cat_mes[chave].get((a, m), 0.0) + _f(ln.valor), 2)

            # saídas por CNPJ+grupo (direto dos lançamentos — pega folha sem banco)
            sai_por_emp: dict[str, dict] = {}
            FinExportService._coletar_saidas(db, a, m, rot, cnpj_limpo, saidas_rows, cat_mes, sai_por_emp)
            # transferências por CNPJ (recebida/enviada)
            transf_por_emp: dict[str, dict] = {}
            FinExportService._coletar_transf(db, a, m, rot, cnpj_limpo, transf_rows, cat_mes, transf_por_emp)

            empresas = set(base) | set(ent_por_emp) | set(sai_por_emp) | set(transf_por_emp)
            for emp in sorted(e for e in empresas if e in ("CMPORT", "TEC")):
                if cnpj_limpo and EMPRESA_POR_CNPJ.get(cnpj_limpo) != emp:
                    continue
                b = base.get(emp, {})
                e = ent_por_emp.get(emp, {})
                s = sai_por_emp.get(emp, {})
                t = transf_por_emp.get(emp, {})
                ent_boleto, ent_recibo = round(e.get("boleto", 0.0), 2), round(e.get("recibo", 0.0), 2)
                s_forn = round(s.get("forn", 0.0), 2); s_desp = round(s.get("desp", 0.0), 2)
                s_func = round(s.get("func", 0.0), 2); s_tar = round(s.get("tar", 0.0), 2)
                rendimento = round(b.get("rendimento", 0.0), 2)
                transf_rec = round(t.get("rec", 0.0), 2); transf_env = round(t.get("env", 0.0), 2)
                saldo_ini = round(b.get("saldo_inicial", 0.0), 2)
                ent_tot = round(ent_boleto + ent_recibo, 2)
                sai_tot = round(s_forn + s_desp + s_func + s_tar, 2)
                saldo_mes = round(saldo_ini + ent_tot + rendimento + transf_rec - transf_env - sai_tot, 2)
                resumo_rows.append({
                    "mes": rot, "empresa": emp, "saldo_inicial": saldo_ini,
                    "ent_boleto": ent_boleto, "ent_recibo": ent_recibo, "ent_avulso": 0.0,
                    "ent_tot": ent_tot, "transf_rec": transf_rec, "transf_env": transf_env,
                    "rendimento": rendimento,
                    "s_forn": s_forn, "s_desp": s_desp, "s_func": s_func, "s_tar": s_tar,
                    "sai_tot": sai_tot, "saldo_mes": saldo_mes,
                    "saldo_extrato": round(b.get("saldo_extrato", 0.0), 2) if b.get("tem_extrato") else None,
                })

        # ── aba Resumo ─────────────────────────────────────────────────────
        ws = wb.create_sheet("Resumo")
        _cab(ws, [
            "Mês", "Empresa", "Saldo inicial",
            "Entradas — Boletos/Notas", "Entradas — Recibos", "Entradas (total)",
            "Transf. recebidas", "Transf. enviadas", "Rendimento",
            "Saídas — Fornecedores", "Saídas — Despesas", "Saídas — Funcionário", "Saídas — Tarifas/IR",
            "Saídas (total)", "Saldo do mês", "Saldo do extrato",
        ])
        campos = ["saldo_inicial", "ent_boleto", "ent_recibo", "ent_tot",
                  "transf_rec", "transf_env", "rendimento",
                  "s_forn", "s_desp", "s_func", "s_tar", "sai_tot", "saldo_mes"]
        for r in resumo_rows:
            ws.append([r["mes"], r["empresa"]] + [round(r[c], 2) for c in campos]
                      + [r["saldo_extrato"]])
        if resumo_rows:
            tot = ["TOTAL", ""] + [round(sum(r[c] for r in resumo_rows), 2) for c in campos] + [""]
            ws.append(tot)
            for cell in ws[ws.max_row]:
                cell.font = _TOT_FONT
        _ajustar_largura(ws, com_filtro=False)

        # ── aba Entradas ──────────────────────────────────────────────────
        ws = wb.create_sheet("Entradas")
        _cab(ws, ["Mês", "Empresa", "Cliente/Condomínio", "Tipo", "Nota/Recibo", "Origem",
                  "Pagamento", "Valor", "Banco", "Obs"])
        for row in sorted(entradas_rows, key=lambda x: (x[0], x[1], x[6])):
            ws.append(row)
        _ajustar_largura(ws)

        # ── aba Saídas ────────────────────────────────────────────────────
        ws = wb.create_sheet("Saídas")
        _cab(ws, ["Mês", "Empresa", "Grupo", "Categoria", "Descrição", "Fornecedor/Funcionário",
                  "Pagamento", "Valor", "Banco", "Forma"])
        for row in sorted(saidas_rows, key=lambda x: (x[0], x[1], x[2], x[6])):
            ws.append(row)
        _ajustar_largura(ws)

        # ── aba Transferências ────────────────────────────────────────────
        ws = wb.create_sheet("Transferências")
        _cab(ws, ["Mês", "Data", "Valor", "De (conta)", "Para (conta)", "Categoria", "Descrição"])
        for row in sorted(transf_rows, key=lambda x: (x[0], x[1])):
            ws.append(row)
        _ajustar_largura(ws)

        # ── aba Categoria x Mês ───────────────────────────────────────────
        ws = wb.create_sheet("Categoria x Mês")
        _cab(ws, ["Categoria"] + [rotulo_mes[k] for k in meses] + ["Total"])
        for cat in sorted(cat_mes.keys()):
            linha = [cat]
            tot = 0.0
            for k in meses:
                v = round(cat_mes[cat].get(k, 0.0), 2)
                linha.append(v)
                tot += v
            linha.append(round(tot, 2))
            ws.append(linha)
        _ajustar_largura(ws)

        # ── aba Pendências (opcional) ─────────────────────────────────────
        if incluir_pendentes:
            ws = wb.create_sheet("Pendências")
            _cab(ws, ["Mês", "Empresa", "Tipo", "Cliente/Descrição", "Nota/Categoria",
                      "Vencimento", "Valor pendente", "Situação"])
            for a, m in meses:
                pend = FluxoFinanceiroService.pendencias_ate_mes(db, a, m, cnpj)
                for ln in pend.linhas:
                    if ln.situacao == "PAGO":
                        continue
                    ws.append([
                        rotulo_mes[(a, m)], ln.empresa or "", f"Entrada — {ln.tipo}",
                        ln.condominio_nome, ln.numero_nota or "",
                        ln.data_vencimento.strftime("%d/%m/%Y") if ln.data_vencimento else "",
                        _f(ln.valor_pendente), ln.situacao,
                    ])
                FinExportService._pendencias_saida(db, a, m, rotulo_mes[(a, m)], cnpj_limpo, ws)
            _ajustar_largura(ws)

        out = io.BytesIO()
        wb.save(out)
        return out.getvalue()

    # ──────────────────────────────────────────────────────────────────────
    @staticmethod
    def _coletar_saidas(db, ano, mes, rot, cnpj_limpo, saidas_rows, cat_mes, por_emp=None):
        dp = aliased(DespesaParcela)
        d = aliased(Despesa)
        fn = aliased(Funcionario)
        b = aliased(Banco)
        c = aliased(CategoriaFinanceira)
        f = aliased(Condominio)
        q = (
            db.query(MovimentacaoFinanceira, c, f, b, d, fn)
            .outerjoin(c, c.id == MovimentacaoFinanceira.categoria_id)
            .outerjoin(f, f.id == MovimentacaoFinanceira.fornecedor_id)
            .outerjoin(b, b.id == MovimentacaoFinanceira.banco_id)
            .outerjoin(dp, dp.movimentacao_id == MovimentacaoFinanceira.id)
            .outerjoin(d, d.id == dp.despesa_id)
            .outerjoin(fn, fn.id == d.funcionario_id)
            .filter(
                MovimentacaoFinanceira.tipo == "SAIDA",
                MovimentacaoFinanceira.deletado_em.is_(None),
                func.year(MovimentacaoFinanceira.data) == ano,
                func.month(MovimentacaoFinanceira.data) == mes,
            )
        )
        for mov, cat, forn, banco, desp, func_ in q.all():
            cnpj_mov = _sd(banco.cnpj_titular) if banco else (_sd(desp.cnpj) if desp else "")
            if cnpj_limpo and cnpj_mov != cnpj_limpo:
                continue
            grupo = cat.grupo if cat else "?"
            nome_cat = cat.nome if cat else "-"
            nome_l = nome_cat.lower()
            # funcionario_id na despesa manda mais que o grupo da categoria da mov
            # (a folha migrada às vezes tem a categoria certa na despesa mas não na mov)
            if func_ is not None or grupo == GrupoCategoria.FUNCIONARIO.value:
                g_label, bucket = "Funcionário", "func"
            elif grupo == GrupoCategoria.FORNECEDOR.value or forn is not None:
                g_label, bucket = "Fornecedor", "forn"
            elif any(t in nome_l for t in _TERMOS_TARIFA):
                g_label, bucket = "Tarifa/IR", "tar"
            else:
                g_label, bucket = "Despesa", "desp"
            emp = EMPRESA_POR_CNPJ.get(cnpj_mov, cnpj_mov or "?")
            saidas_rows.append([
                rot, emp, g_label, nome_cat, (mov.descricao or "")[:120],
                (func_.nome if func_ else (forn.nome if forn else "")),
                mov.data.strftime("%d/%m/%Y") if mov.data else "",
                _f(mov.valor), banco.nome if banco else "", mov.forma_pagamento or "",
            ])
            chave = f"Saída — {nome_cat}"
            cat_mes.setdefault(chave, {})
            cat_mes[chave][(ano, mes)] = round(cat_mes[chave].get((ano, mes), 0.0) + _f(mov.valor), 2)
            if por_emp is not None:
                por_emp.setdefault(emp, {"forn": 0.0, "desp": 0.0, "func": 0.0, "tar": 0.0})
                por_emp[emp][bucket] += _f(mov.valor)

    @staticmethod
    def _coletar_transf(db, ano, mes, rot, cnpj_limpo, transf_rows, cat_mes, por_emp=None):
        bo = aliased(Banco)
        bd = aliased(Banco)
        c = aliased(CategoriaFinanceira)
        q = (
            db.query(MovimentacaoFinanceira, bo, bd, c)
            .outerjoin(bo, bo.id == MovimentacaoFinanceira.banco_origem_id)
            .outerjoin(bd, bd.id == MovimentacaoFinanceira.banco_id)
            .outerjoin(c, c.id == MovimentacaoFinanceira.categoria_id)
            .filter(
                MovimentacaoFinanceira.tipo == "ENTRADA",
                MovimentacaoFinanceira.deletado_em.is_(None),
                MovimentacaoFinanceira.banco_origem_id.isnot(None),
                func.year(MovimentacaoFinanceira.data) == ano,
                func.month(MovimentacaoFinanceira.data) == mes,
            )
        )
        for mov, borig, bdest, cat in q.all():
            emp_orig = EMPRESA_POR_CNPJ.get(_sd(borig.cnpj_titular)) if borig else None
            emp_dest = EMPRESA_POR_CNPJ.get(_sd(bdest.cnpj_titular)) if bdest else None
            if cnpj_limpo:
                alvo = EMPRESA_POR_CNPJ.get(cnpj_limpo)
                if alvo not in (emp_orig, emp_dest):
                    continue
            transf_rows.append([
                rot, mov.data.strftime("%d/%m/%Y") if mov.data else "", _f(mov.valor),
                borig.nome + (f" ({borig.razao_social_titular})" if borig and borig.razao_social_titular else "") if borig else "?",
                bdest.nome + (f" ({bdest.razao_social_titular})" if bdest and bdest.razao_social_titular else "") if bdest else "?",
                cat.nome if cat else "", (mov.descricao or "")[:80],
            ])
            cat_mes.setdefault("Transferência interna", {})
            cat_mes["Transferência interna"][(ano, mes)] = round(
                cat_mes["Transferência interna"].get((ano, mes), 0.0) + _f(mov.valor), 2)
            if por_emp is not None:
                if emp_dest:
                    por_emp.setdefault(emp_dest, {"rec": 0.0, "env": 0.0})["rec"] += _f(mov.valor)
                if emp_orig:
                    por_emp.setdefault(emp_orig, {"rec": 0.0, "env": 0.0})["env"] += _f(mov.valor)

    @staticmethod
    def _pendencias_saida(db, ano, mes, rot, cnpj_limpo, ws):
        ultimo = date(ano, mes, calendar.monthrange(ano, mes)[1])
        primeiro = date(ano, mes, 1)
        q = (
            db.query(DespesaParcela, Despesa, CategoriaFinanceira)
            .join(Despesa, Despesa.id == DespesaParcela.despesa_id)
            .outerjoin(CategoriaFinanceira, CategoriaFinanceira.id == Despesa.categoria_id)
            .filter(
                Despesa.deletado_em.is_(None),
                DespesaParcela.status == StatusParcelaDespesa.PENDENTE,
                DespesaParcela.data_vencimento.between(primeiro, ultimo),
            )
        )
        for parc, desp, cat in q.all():
            if cnpj_limpo and _sd(desp.cnpj) != cnpj_limpo:
                continue
            ws.append([
                rot, EMPRESA_POR_CNPJ.get(_sd(desp.cnpj), desp.cnpj), "Saída — despesa/parcela",
                desp.descricao, cat.nome if cat else "",
                parc.data_vencimento.strftime("%d/%m/%Y") if parc.data_vencimento else "",
                _f(parc.valor), "PENDENTE",
            ])
