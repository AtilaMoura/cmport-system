from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile, File
from fastapi.responses import Response
from sqlalchemy.orm import Session
from typing import List, Optional
from io import BytesIO
from datetime import datetime
from pydantic import BaseModel

from app.core.database import SessionLocal
from app.core.dependencies import get_storage_client, require_admin
from app.core.storage_client import StorageClient
from app.services.nota_fiscal_service import NotaFiscalService, corrigir_datas_servico
from app.schemas.nota_fiscal_schema import (
    NotaFiscalCreate, NotaFiscalResponse, ImportacaoResponse, NotaFiscalUpdate,
    VincularNotasRequest, CandidataVinculoResponse, UploadPdfResponse,
)


class VincularCondominioRequest(BaseModel):
    condominio_id: int

router = APIRouter()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.post("", response_model=NotaFiscalResponse, status_code=201)
def create_nota(nota: NotaFiscalCreate, db: Session = Depends(get_db)):
    try:
        return NotaFiscalService.create_nota(db, nota)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao criar nota: {str(e)}")


@router.get("", response_model=List[NotaFiscalResponse])
def list_notas(db: Session = Depends(get_db)):
    return NotaFiscalService.get_all_notas(db)


@router.get("/exportar")
def exportar_notas_excel(
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    condominio_id: Optional[int] = Query(None),
    tipo: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    from datetime import date as date_type
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from app.models.nota_fiscal_model import NotaFiscal, StatusNota
    from app.models.condominio_model import Condominio

    query = db.query(NotaFiscal).join(
        Condominio, NotaFiscal.condominio_id == Condominio.id, isouter=True
    ).filter(NotaFiscal.status == StatusNota.AUTORIZADA)

    if data_inicio:
        query = query.filter(NotaFiscal.data_vencimento >= date_type.fromisoformat(data_inicio))
    if data_fim:
        query = query.filter(NotaFiscal.data_vencimento <= date_type.fromisoformat(data_fim))
    if condominio_id:
        query = query.filter(NotaFiscal.condominio_id == condominio_id)
    if tipo:
        query = query.filter(NotaFiscal.tipo == tipo)

    notas = query.order_by(NotaFiscal.data_vencimento.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['ID', 'Número', 'Tipo', 'Status', 'Condomínio', 'Valor', 'Parcelas', 'Vencimento', 'Pagamento', 'Cliente', 'Observação']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for nota in notas:
        cond_nome = nota.condominio.nome if nota.condominio else "Sem condomínio"
        ws.append([
            nota.id, nota.numero_nota, nota.tipo.value, nota.status.value, cond_nome,
            float(nota.valor), nota.parcelas,
            nota.data_vencimento.strftime('%d/%m/%Y') if nota.data_vencimento else '',
            nota.data_pagamento.strftime('%d/%m/%Y') if nota.data_pagamento else '',
            nota.cliente_nome or '', nota.observacao or '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 2, 50
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/vincular-notas")
def vincular_notas(body: VincularNotasRequest, db: Session = Depends(get_db)):
    """Vincula duas notas do mesmo condomínio: deleta os dois serviços e cria um serviço combinado."""
    try:
        return NotaFiscalService.vincular_notas(db, body.nota_a_id, body.nota_b_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/{id}/desvincular-notas")
def desvincular_notas(id: int, db: Session = Depends(get_db)):
    """Desfaz o vínculo entre duas notas e recria os dois serviços individuais."""
    try:
        return NotaFiscalService.desvincular_notas(db, id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/candidatas-vinculo/{servico_id}", response_model=List[CandidataVinculoResponse])
def get_candidatas_vinculo(servico_id: int, db: Session = Depends(get_db)):
    """Retorna notas do mesmo condomínio que podem ser vinculadas ao serviço informado."""
    try:
        return NotaFiscalService.get_candidatas_vinculo(db, servico_id)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/{id}", response_model=NotaFiscalResponse)
def get_nota_by_id(id: int, db: Session = Depends(get_db)):
    nota = NotaFiscalService.get_nota_by_id(db, id)
    if not nota:
        raise HTTPException(status_code=404, detail="Nota Fiscal não encontrada")
    return nota


@router.get("/numero/{numero}", response_model=NotaFiscalResponse)
def get_nota_by_numero(numero: str, db: Session = Depends(get_db)):
    nota = NotaFiscalService.get_nota_by_numero(db, numero)
    if not nota:
        raise HTTPException(status_code=404, detail="Nota Fiscal não encontrada")
    return nota


@router.post("/importar-xml", response_model=ImportacaoResponse)
async def importar_xmls(
    arquivos: List[UploadFile] = File(...),
    tipo: Optional[str] = Form(None),
    db: Session = Depends(get_db),
    storage: StorageClient = Depends(get_storage_client)
):
    try:
        resultado = await NotaFiscalService.importar_xmls(db, arquivos, tipo, storage)
        return ImportacaoResponse(
            processados=resultado["processados"],
            ja_existentes=resultado.get("ja_existentes", 0),
            canceladas=resultado.get("canceladas", 0),
            erros=resultado["erros"]
        )
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise HTTPException(status_code=500, detail=f"Erro ao importar XMLs: {str(e)}")


@router.put("/{id}", response_model=NotaFiscalResponse)
def update_nota(id: int, nota_update: NotaFiscalUpdate, db: Session = Depends(get_db)):
    updated = NotaFiscalService.update_nota(db, id, nota_update)
    if not updated:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    return updated


@router.patch("/{id}/vincular-condominio", response_model=NotaFiscalResponse)
def vincular_condominio(id: int, body: VincularCondominioRequest, db: Session = Depends(get_db)):
    """Vincula (ou re-vincula) uma nota fiscal a um condomínio."""
    try:
        nota, aviso = NotaFiscalService.vincular_condominio(db, id, body.condominio_id)
        if aviso:
            print(f"[VincularCondominio] nota {id}: {aviso}")
        return nota
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.patch("/{id}/dispensar-alerta", response_model=NotaFiscalResponse)
def dispensar_alerta_impostos(id: int, db: Session = Depends(get_db)):
    """Remove o alerta de divergência de impostos de uma nota fiscal."""
    from app.models.nota_fiscal_model import NotaFiscal
    db_nota = db.query(NotaFiscal).filter(NotaFiscal.id == id).first()
    if not db_nota:
        raise HTTPException(status_code=404, detail="Nota não encontrada.")
    db_nota.alerta_impostos = 0
    db_nota.divergencia_impostos = None
    db.commit()
    db.refresh(db_nota)
    return db_nota


@router.post("/{id}/revalidar")
def revalidar_nota(id: int, db: Session = Depends(get_db)):
    """Re-parseia o xml_original e corrige o status da nota."""
    return NotaFiscalService.revalidar_nota_do_xml(db, id)


@router.post("/{id}/revalidar-completo")
def revalidar_nota_completo(id: int, db: Session = Depends(get_db)):
    """Re-parseia o xml_original e atualiza TODOS os campos: tipo, vencimento, parcelas, impostos."""
    return NotaFiscalService.revalidar_campos_do_xml(db, id)


@router.post("/revalidar-todas")
def revalidar_todas_notas(db: Session = Depends(get_db)):
    """Re-parseia o XML de todas as notas e corrige status incorretos."""
    return NotaFiscalService.revalidar_todas(db)


@router.post("/corrigir-datas-servico")
def corrigir_datas(db: Session = Depends(get_db)):
    """Re-parseia o XML de todas as notas e corrige data_servico nos registros de ManutencaoAssistencia."""
    return corrigir_datas_servico(db)


@router.post("/revalidar-todas-completo")
def revalidar_todas_completo(db: Session = Depends(get_db)):
    """Re-parseia o XML de todas as notas e atualiza TODOS os campos (tipo, vencimento, parcelas, impostos)."""
    from app.models.nota_fiscal_model import NotaFiscal
    notas = db.query(NotaFiscal).filter(NotaFiscal.xml_original.isnot(None)).all()
    total = len(notas)
    alteradas = 0
    erros = 0
    detalhes = []
    for nota in notas:
        r = NotaFiscalService.revalidar_campos_do_xml(db, nota.id)
        if r.get("alteracoes"):
            alteradas += 1
            detalhes.append({"nota_id": r["nota_id"], "numero": r.get("numero_nota"), "alteracoes": r["alteracoes"]})
        if r.get("resultado") in ("erro", "tipo_nao_suportado"):
            erros += 1
    return {
        "total": total,
        "alteradas": alteradas,
        "erros": erros,
        "detalhes": detalhes,
        "mensagem": f"Revalidação completa: {total} notas, {alteradas} atualizadas, {erros} erros.",
    }


@router.post("/{id}/pdf", response_model=UploadPdfResponse)
async def upload_pdf_nota(
    id: int,
    pdf: UploadFile = File(...),
    db: Session = Depends(get_db),
    storage: StorageClient = Depends(get_storage_client)
):
    """Faz upload manual de um PDF para uma nota fiscal existente."""
    if not pdf.filename.lower().endswith('.pdf'):
        raise HTTPException(status_code=422, detail="O arquivo deve ser um PDF.")
    
    conteudo = await pdf.read()
    object_key = NotaFiscalService.upload_pdf_nota(db, id, conteudo, storage)
    
    return UploadPdfResponse(
        nota_id=id,
        pdf_object_key=object_key,
        mensagem="PDF enviado com sucesso."
    )


@router.get("/{id}/pdf-url")
def get_pdf_url(
    id: int,
    db: Session = Depends(get_db),
    storage: StorageClient = Depends(get_storage_client)
):
    """Retorna uma URL assinada (temporária) para visualizar o PDF da nota."""
    return NotaFiscalService.get_pdf_url(db, id, storage)


@router.delete("/{id}/pdf", status_code=204)
def delete_pdf_nota(
    id: int,
    db: Session = Depends(get_db),
    storage: StorageClient = Depends(get_storage_client),
    admin = Depends(require_admin)
):
    """Remove o PDF da nota do storage e limpa a referência no banco."""
    NotaFiscalService.delete_pdf_nota(db, id, storage)
    return Response(status_code=204)


@router.delete("/{id}", status_code=204)
def delete_nota(
    id: int,
    motivo: str = Query(None),
    deletar_servicos: bool = Query(False),
    db: Session = Depends(get_db),
    storage: StorageClient = Depends(get_storage_client)
):
    if not NotaFiscalService.delete_nota(db=db, nota_id=id, motivo=motivo, deletar_servicos=deletar_servicos, storage=storage):
        raise HTTPException(status_code=404, detail="Nota não encontrada")
