from fastapi import APIRouter, Depends, Query, Response
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import datetime, date, timedelta
from typing import Optional, List
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.core.database import SessionLocal
from app.models.nota_fiscal_model import NotaFiscal, StatusNota
from app.models.servico_model import ManutencaoAssistencia
from app.models.condominio_model import Condominio
from app.repositories.boleto_repository import BoletoRepository

router = APIRouter()


def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.get("/servicos/exportar")
def exportar_servicos_excel(
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    condominio_id: Optional[int] = Query(None),
    tipo: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    from datetime import date as date_type

    query = db.query(ManutencaoAssistencia).join(
        Condominio, ManutencaoAssistencia.condominio_id == Condominio.id, isouter=True
    )
    if data_inicio:
        query = query.filter(ManutencaoAssistencia.data_servico >= date_type.fromisoformat(data_inicio))
    if data_fim:
        query = query.filter(ManutencaoAssistencia.data_servico <= date_type.fromisoformat(data_fim))
    if condominio_id:
        query = query.filter(ManutencaoAssistencia.condominio_id == condominio_id)
    if tipo:
        query = query.filter(ManutencaoAssistencia.tipo == tipo)

    servicos = query.order_by(ManutencaoAssistencia.data_servico.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Servicos"
    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['ID', 'Tipo', 'Data Servico', 'Condominio', 'Nota Fiscal ID', 'Descricao', 'Criado Em']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for s in servicos:
        cond_nome = s.condominio.nome if s.condominio else "Sem condominio"
        ws.append([
            s.id,
            s.tipo.value if hasattr(s.tipo, 'value') else str(s.tipo),
            s.data_servico.strftime('%d/%m/%Y') if s.data_servico else '',
            cond_nome,
            s.nota_fiscal_id or '',
            (s.descricao or '')[:200],
            s.criado_em.strftime('%d/%m/%Y %H:%M') if s.criado_em else '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 2, 60
        )

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"servicos_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/estatisticas")
def get_dashboard_stats(db: Session = Depends(get_db)):
    hoje = date.today()
    ano_atual = hoje.year

    primeiro_dia_mes_atual = hoje.replace(day=1)
    ultimo_dia_mes_passado = primeiro_dia_mes_atual - timedelta(days=1)
    mes_passado = ultimo_dia_mes_passado.month
    ano_mes_passado = ultimo_dia_mes_passado.year
    ano_passado = ano_atual - 1

    nf_ativas = NotaFiscal.status == StatusNota.AUTORIZADA

    stats_notas_mes_atual = db.query(
        func.count(NotaFiscal.id), func.sum(NotaFiscal.valor)
    ).filter(
        extract('month', NotaFiscal.data_vencimento) == hoje.month,
        extract('year', NotaFiscal.data_vencimento) == ano_atual,
        nf_ativas,
    ).first()
    notas_mes_atual = stats_notas_mes_atual[0] or 0
    valor_mes_atual = stats_notas_mes_atual[1] or 0.0

    stats_notas_mes_passado = db.query(
        func.count(NotaFiscal.id), func.sum(NotaFiscal.valor)
    ).filter(
        extract('month', NotaFiscal.data_vencimento) == mes_passado,
        extract('year', NotaFiscal.data_vencimento) == ano_mes_passado,
        nf_ativas,
    ).first()
    notas_mes_passado = stats_notas_mes_passado[0] or 0
    valor_mes_passado = stats_notas_mes_passado[1] or 0.0

    valor_ano_atual = db.query(func.sum(NotaFiscal.valor)).filter(
        extract('year', NotaFiscal.data_vencimento) == ano_atual, nf_ativas,
    ).scalar() or 0.0

    valor_ano_passado = db.query(func.sum(NotaFiscal.valor)).filter(
        extract('year', NotaFiscal.data_vencimento) == ano_passado, nf_ativas,
    ).scalar() or 0.0

    total_notas_geral = db.query(func.count(NotaFiscal.id)).filter(nf_ativas).scalar() or 0
    total_valor_geral = db.query(func.sum(NotaFiscal.valor)).filter(nf_ativas).scalar() or 0.0

    servicos_mes_atual = db.query(func.count(ManutencaoAssistencia.id)).filter(
        extract('month', ManutencaoAssistencia.data_servico) == hoje.month,
        extract('year', ManutencaoAssistencia.data_servico) == ano_atual
    ).scalar() or 0

    servicos_mes_passado = db.query(func.count(ManutencaoAssistencia.id)).filter(
        extract('month', ManutencaoAssistencia.data_servico) == mes_passado,
        extract('year', ManutencaoAssistencia.data_servico) == ano_mes_passado
    ).scalar() or 0

    stats_servicos_tipo = db.query(
        ManutencaoAssistencia.tipo, func.count(ManutencaoAssistencia.id)
    ).group_by(ManutencaoAssistencia.tipo).all()

    manutencoes = next((c for t, c in stats_servicos_tipo if t == 'manutencao'), 0)
    assistencias = next((c for t, c in stats_servicos_tipo if t == 'assistencia'), 0)
    total_servicos = manutencoes + assistencias

    top_condominios = db.query(
        Condominio.nome, func.sum(NotaFiscal.valor).label('total')
    ).join(NotaFiscal, NotaFiscal.condominio_id == Condominio.id
    ).filter(nf_ativas
    ).group_by(Condominio.id, Condominio.nome
    ).order_by(func.sum(NotaFiscal.valor).desc()
    ).limit(5).all()

    servicos_por_mes = db.query(
        extract('month', ManutencaoAssistencia.data_servico).label('mes'),
        extract('year', ManutencaoAssistencia.data_servico).label('ano'),
        func.count(ManutencaoAssistencia.id).label('quantidade')
    ).filter(ManutencaoAssistencia.data_servico >= hoje - timedelta(days=365)
    ).group_by('mes', 'ano'
    ).order_by(func.count(ManutencaoAssistencia.id).desc()
    ).limit(5).all()

    servicos_por_dia_semana = db.query(
        func.dayofweek(ManutencaoAssistencia.data_servico).label('dia_semana_mysql'),
        func.count(ManutencaoAssistencia.id).label('quantidade')
    ).group_by('dia_semana_mysql').all()

    dias_semana_nomes = ['Domingo', 'Segunda', 'Terça', 'Quarta', 'Quinta', 'Sexta', 'Sábado']
    ranking_dias_dict = {dia: 0 for dia in dias_semana_nomes}
    for r in servicos_por_dia_semana:
        if r.dia_semana_mysql is not None:
            dia_index = int(r.dia_semana_mysql) - 1
            ranking_dias_dict[dias_semana_nomes[dia_index]] = int(r.quantidade or 0)

    def calcular_variacao(atual, passado):
        if passado is None or passado == 0:
            return 100.0 if atual > 0 else 0.0
        return round(((atual - passado) / passado) * 100, 2)

    boleto_stats = BoletoRepository.get_stats(db)

    return {
        "resumo_geral": {
            "total_condominios": db.query(func.count(Condominio.id)).scalar() or 0,
            "total_notas": total_notas_geral,
            "total_valor_notas": float(total_valor_geral),
            "total_servicos": total_servicos,
            "total_manutencoes": manutencoes,
            "total_assistencias": assistencias,
        },
        "mes_atual": {
            "notas": notas_mes_atual,
            "valor": float(valor_mes_atual),
            "servicos": servicos_mes_atual,
            "variacao_notas_percentual": calcular_variacao(notas_mes_atual, notas_mes_passado),
            "variacao_valor_percentual": calcular_variacao(valor_mes_atual, valor_mes_passado),
            "variacao_servicos_percentual": calcular_variacao(servicos_mes_atual, servicos_mes_passado),
        },
        "mes_passado": {
            "notas": notas_mes_passado,
            "valor": float(valor_mes_passado),
            "servicos": servicos_mes_passado,
        },
        "ano_atual": {
            "valor": float(valor_ano_atual),
            "variacao_percentual": calcular_variacao(valor_ano_atual, valor_ano_passado),
        },
        "ano_passado": {"valor": float(valor_ano_passado)},
        "rankings": {
            "top_condominios": [{"nome": nome, "valor": float(total or 0)} for nome, total in top_condominios],
            "top_meses_servicos": [
                {"mes": int(mes), "ano": int(ano), "quantidade": int(qtd),
                 "nome_mes": ['Jan','Fev','Mar','Abr','Mai','Jun','Jul','Ago','Set','Out','Nov','Dez'][int(mes)-1]}
                for mes, ano, qtd in servicos_por_mes
            ],
            "dias_semana": [{"dia": dia, "quantidade": qtd} for dia, qtd in ranking_dias_dict.items()],
        },
        "boletos": boleto_stats,
    }
