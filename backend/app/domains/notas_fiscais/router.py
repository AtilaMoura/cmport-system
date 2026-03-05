from fastapi import APIRouter, Depends, Form, HTTPException, Query, UploadFile, File
from sqlalchemy.orm import Session
from typing import List, Optional

from app.core.database import SessionLocal
from .service import NotaFiscalService
from .schema import NotaFiscalCreate, NotaFiscalResponse, ImportacaoResponse, NotaFiscalUpdate


router = APIRouter()


# Dependência para obter a sessão do banco
def get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


@router.post("/", response_model=NotaFiscalResponse, status_code=201)
def create_nota(nota: NotaFiscalCreate, db: Session = Depends(get_db)):
    """
    Cria uma nova Nota Fiscal manualmente no sistema
    """
    try:
        return NotaFiscalService.create_nota(db, nota)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Erro ao criar nota: {str(e)}")


@router.get("/", response_model=List[NotaFiscalResponse])
def list_notas(db: Session = Depends(get_db)):
    """
    Lista todas as notas fiscais cadastradas no sistema
    """
    return NotaFiscalService.get_all_notas(db)


@router.get("/exportar")
def exportar_notas_excel(
    data_inicio: Optional[str] = Query(None),
    data_fim: Optional[str] = Query(None),
    condominio_id: Optional[int] = Query(None),
    tipo: Optional[str] = Query(None),
    db: Session = Depends(get_db)
):
    """Exporta notas fiscais para Excel."""
    from datetime import date as date_type, datetime
    from io import BytesIO
    from fastapi.responses import Response
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from .model import NotaFiscal, StatusNota
    from app.domains.condominios.model import Condominio

    query = db.query(NotaFiscal).join(
        Condominio, NotaFiscal.condominio_id == Condominio.id, isouter=True
    ).filter(NotaFiscal.status != StatusNota.CANCELADA)

    if data_inicio:
        query = query.filter(NotaFiscal.data_vencimento >= date_type.fromisoformat(data_inicio))
    if data_fim:
        query = query.filter(NotaFiscal.data_vencimento <= date_type.fromisoformat(data_fim))
    if condominio_id:
        query = query.filter(NotaFiscal.condominio_id == condominio_id)
    if tipo:
        query = query.filter(NotaFiscal.tipo == tipo)

    notas = query.order_by(NotaFiscal.data_vencimento.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Notas Fiscais"

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    headers = ['ID', 'Número', 'Tipo', 'Status', 'Condomínio', 'Valor', 'Parcelas',
               'Vencimento', 'Pagamento', 'Cliente', 'Observação']
    ws.append(headers)
    for col_num, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    for nota in notas:
        cond_nome = nota.condominio.nome if nota.condominio else "Sem condomínio"
        ws.append([
            nota.id, nota.numero_nota, nota.tipo.value, nota.status.value, cond_nome,
            float(nota.valor), nota.parcelas,
            nota.data_vencimento.strftime('%d/%m/%Y') if nota.data_vencimento else '',
            nota.data_pagamento.strftime('%d/%m/%Y') if nota.data_pagamento else '',
            nota.cliente_nome or '', nota.observacao or '',
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = min(
            max(len(str(c.value or '')) for c in col) + 2, 50
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    filename = f"notas_fiscais_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.get("/{id}", response_model=NotaFiscalResponse)
def get_nota_by_id(id: int, db: Session = Depends(get_db)):
    """
    Busca uma nota específica pelo ID
    """
    nota = NotaFiscalService.get_nota_by_id(db, id)
    if not nota:
        raise HTTPException(status_code=404, detail="Nota Fiscal não encontrada")
    return nota


@router.get("/numero/{numero}", response_model=NotaFiscalResponse)
def get_nota_by_numero(numero: str, db: Session = Depends(get_db)):
    """
    Busca uma nota específica pelo número
    """
    nota = NotaFiscalService.get_nota_by_numero(db, numero)
    if not nota:
        raise HTTPException(status_code=404, detail="Nota Fiscal não encontrada")
    return nota


@router.post("/importar-xml", response_model=ImportacaoResponse)
async def importar_xmls(
    arquivos: List[UploadFile] = File(..., description="Arquivos XML ou ZIP contendo XMLs de NFe"),
    tipo: Optional[str] = Form(None),
    db: Session = Depends(get_db)
):
    """
    Importa Notas Fiscais Eletrônicas (NFe) a partir de arquivos XML
    
    **Aceita:**
    - Múltiplos arquivos `.xml` individuais
    - Arquivo `.zip` contendo vários XMLs (inclusive em subpastas)
    - Combinação de ambos
    
    **Funcionalidades:**
    - Extrai dados principais da NFe (número, série, emitente, destinatário, valor)
    - Vincula automaticamente ao condomínio pelo CNPJ (se encontrado)
    - Detecta tipo de serviço (Manutenção/Assistência) pelas informações complementares
    - Identifica número de parcelas
    - Evita duplicidade pela chave de acesso
    - Armazena XML completo para auditoria
    
    **Retorna:**
    - Quantidade de notas processadas com sucesso
    - Lista de erros (se houver)
    """
    try:
        resultado = await NotaFiscalService.importar_xmls(db, arquivos, tipo)
        return ImportacaoResponse(
            processados=resultado["processados"],
            canceladas=resultado.get("canceladas", 0),  # ✅ fix
            erros=resultado["erros"]
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"Erro ao importar XMLs: {str(e)}"
        )

@router.put("/{id}", response_model=NotaFiscalResponse)
def update_nota(id: int, nota_update: NotaFiscalUpdate, db: Session = Depends(get_db)):
    updated = NotaFiscalService.update_nota(db, id, nota_update)
    if not updated:
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    return updated


@router.delete("/{id}", status_code=204)
def delete_nota(
    id: int,
    motivo: str = Query(None, description="Motivo da exclusão (opcional)"),
    deletar_servicos: bool = Query(
        False, 
        description="Se true, deleta também todos os serviços vinculados à nota"
    ),
    db: Session = Depends(get_db)
):
    if not NotaFiscalService.delete_nota(
        db=db, 
        nota_id=id, 
        motivo=motivo, 
        deletar_servicos=deletar_servicos
    ):
        raise HTTPException(status_code=404, detail="Nota não encontrada")
    return None