"""
Relatório comparativo: NFs na planilha x NFs exportadas do sistema
Fontes:
  - Entradas Fluxo Janeiro.xlsx  (planilha financeira)
  - notas_fiscais_2026-05-27.xlsx (export do sistema)
"""
import re
import sys
from pathlib import Path
from datetime import datetime

try:
    import openpyxl
except ImportError:
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"], check=True)
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = Path(__file__).parent.parent.parent

MESES = {
    1: "Janeiro", 2: "Fevereiro", 3: "Marco", 4: "Abril",
    5: "Maio", 6: "Junho", 7: "Julho", 8: "Agosto",
    9: "Setembro", 10: "Outubro", 11: "Novembro", 12: "Dezembro",
}

# ─────────────────────────────────────────────
# 1. LER PLANILHA FINANCEIRA
# ─────────────────────────────────────────────
PLANILHA_PATH = BASE / "Entradas Fluxo Janeiro.xlsx"
print(f"Lendo planilha: {PLANILHA_PATH}")

wb_plan = openpyxl.load_workbook(PLANILHA_PATH, data_only=True, read_only=True)
ws_plan = wb_plan.active

planilha_nfs = {}  # {mes_num: [dict]}
mes_atual = None

for row in ws_plan.iter_rows(min_row=1, max_col=12, values_only=True):
    cod      = row[0]   # Col A — código como "Contrato12026"
    cond     = row[2]   # Col C
    categ    = row[3]   # Col D
    nf_raw   = row[4]   # Col E
    parcela  = row[5]   # Col F
    vencto   = row[7]   # Col H
    valor    = row[10]  # Col K

    # Detecta mês pelo código da linha
    if cod and isinstance(cod, str):
        m = re.search(r"(\d+)2026", cod)
        if m:
            mes_atual = int(m.group(1))

    if not nf_raw:
        continue
    try:
        nf_num = int(nf_raw)
    except (ValueError, TypeError):
        continue
    if nf_num < 1000:
        continue
    if mes_atual is None:
        continue

    venc_str = vencto.strftime("%d/%m/%Y") if hasattr(vencto, "strftime") else str(vencto or "")
    planilha_nfs.setdefault(mes_atual, []).append({
        "numero_nota": nf_num,
        "condominio": str(cond).strip() if cond else "",
        "categoria":  str(categ).strip() if categ else "",
        "parcela":    str(parcela or ""),
        "vencimento": venc_str,
        "valor":      float(valor) if isinstance(valor, (int, float)) else None,
    })

print("NFs na planilha por mes:")
for m, lst in sorted(planilha_nfs.items()):
    nums = sorted(set(e["numero_nota"] for e in lst))
    print(f"  {MESES.get(m, m)}: {len(lst)} NFs  ({min(nums)}-{max(nums)})")

# ─────────────────────────────────────────────
# 2. LER EXPORT DO SISTEMA
# ─────────────────────────────────────────────
SISTEMA_PATH = BASE / "notas_fiscais_2026-05-27.xlsx"
print(f"\nLendo sistema: {SISTEMA_PATH}")

wb_sis = openpyxl.load_workbook(SISTEMA_PATH, data_only=True, read_only=True)
ws_sis = wb_sis.active

# {numero_nota_str: dict} — chave é o campo "Número" original
sistema_por_numero = {}  # chave: numero limpo (apenas dígitos para NFs inteiras)
sistema_lista = []       # todas as NFs do sistema para exibir por mês

for row in ws_sis.iter_rows(min_row=2, max_col=11, values_only=True):
    if not row[0]:
        continue
    id_, num, tipo, status, cond, valor, parcelas, venc, pag, cliente, obs = (row + (None,) * 11)[:11]
    if not num:
        continue
    num_str = str(num).strip()
    venc_str = str(venc or "").strip()

    # Extrai mês do vencimento (formato dd/mm/yyyy)
    mes_venc = None
    m = re.search(r"\d{2}/(\d{2})/\d{4}", venc_str)
    if m:
        mes_venc = int(m.group(1))

    entry = {
        "id": id_,
        "numero_nota": num_str,
        "tipo": str(tipo or "").strip(),
        "status": str(status or "").strip(),
        "condominio": str(cond or "").strip(),
        "valor": float(valor) if isinstance(valor, (int, float)) else None,
        "parcelas": parcelas,
        "vencimento": venc_str,
        "mes_venc": mes_venc,
    }
    sistema_lista.append(entry)
    sistema_por_numero[num_str] = entry

# Para comparação com planilha (que usa só inteiros), mapeamos também por inteiro
sistema_por_int = {}
for num_str, entry in sistema_por_numero.items():
    try:
        sistema_por_int[int(num_str)] = entry
    except (ValueError, TypeError):
        pass  # NFs com formato "117-2" são ignoradas nesta chave

print(f"NFs no sistema: {len(sistema_lista)}")
print("NFs no sistema por mes de vencimento:")
from collections import Counter
contagem = Counter(e["mes_venc"] for e in sistema_lista if e["mes_venc"])
for mes, cnt in sorted(contagem.items()):
    print(f"  {MESES.get(mes, mes)}: {cnt}")

# ─────────────────────────────────────────────
# 3. ESTILOS XLSX
# ─────────────────────────────────────────────
COR_CABEC   = PatternFill("solid", fgColor="1F4E79")
COR_FALTA   = PatternFill("solid", fgColor="FFCCCC")
COR_OK      = PatternFill("solid", fgColor="C6EFCE")
COR_SODB    = PatternFill("solid", fgColor="FFF2CC")
COR_TITULO  = PatternFill("solid", fgColor="D9E1F2")

BORDA = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)

def cel(ws, r, c, value=None, fill=None, bold=False, fmt=None, align="left"):
    cell = ws.cell(row=r, column=c, value=value)
    cell.border = BORDA
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if fill:
        cell.fill = fill
    if bold or fill == COR_CABEC:
        cell.font = Font(bold=True, color="FFFFFF" if fill == COR_CABEC else "000000")
    if fmt:
        cell.number_format = fmt
    return cell

def cabecalho(ws, row, cols):
    for c, titulo in enumerate(cols, 1):
        cell = ws.cell(row=row, column=c, value=titulo)
        cell.fill = COR_CABEC
        cell.font = Font(bold=True, color="FFFFFF")
        cell.border = BORDA
        cell.alignment = Alignment(horizontal="center", vertical="center")

def titulo_secao(ws, row, texto, ncols, fill):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    c = ws.cell(row=row, column=1, value=texto)
    c.fill = fill
    c.font = Font(bold=True, size=11)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 20

def ajustar(ws, larguras):
    for i, w in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def fmt_r(v):
    try:
        return float(v)
    except Exception:
        return None

# ─────────────────────────────────────────────
# 4. GERAR XLSX
# ─────────────────────────────────────────────
wb_out = Workbook()
wb_out.remove(wb_out.active)

resumo_dados = []  # [(mes, total_plan, total_sis, faltando, ok, so_sis)]

NCOLS = 7
COLS_FALTA = ["Nº NF", "Condominio (planilha)", "Categoria", "Parcela", "Vencimento", "Valor (R$)", "Acao"]
COLS_OK    = ["Nº NF", "Condominio (sistema)", "Tipo", "Vencimento", "Valor (R$)", "Status", ""]
COLS_SIS   = ["Nº NF", "Condominio", "Tipo", "Vencimento", "Valor (R$)", "Status", ""]
LARG       = [12, 40, 14, 14, 16, 14, 22]

for mes_num in sorted(planilha_nfs.keys()):
    nome_mes = MESES.get(mes_num, str(mes_num))
    entradas = planilha_nfs[mes_num]
    nfs_plan_nums = sorted(set(e["numero_nota"] for e in entradas))

    # NFs do sistema neste mês
    sis_mes = [e for e in sistema_lista if e["mes_venc"] == mes_num]
    sis_mes_nums = set()
    for e in sis_mes:
        try:
            sis_mes_nums.add(int(e["numero_nota"]))
        except Exception:
            pass

    faltando = [n for n in nfs_plan_nums if n not in sistema_por_int]
    em_ambos = [n for n in nfs_plan_nums if n in sistema_por_int]
    so_sis   = sorted(sis_mes_nums - set(nfs_plan_nums))

    resumo_dados.append((nome_mes, len(nfs_plan_nums), len(sis_mes), len(faltando), len(em_ambos), len(so_sis)))

    ws = wb_out.create_sheet(title=nome_mes)
    row = 1

    # Título da aba
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=NCOLS)
    t = ws.cell(row=row, column=1, value=f"Relatorio NFs — {nome_mes} 2026")
    t.font = Font(bold=True, size=14)
    t.fill = COR_TITULO
    t.alignment = Alignment(horizontal="center")
    ws.row_dimensions[row].height = 24
    row += 1

    ws.cell(row=row, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    row += 1

    # Métricas
    metricas = [
        ("NFs na planilha financeira", len(nfs_plan_nums)),
        (f"NFs no sistema (venc. {nome_mes})", len(sis_mes)),
        ("FALTANDO no sistema", len(faltando)),
        ("Presentes em ambos", len(em_ambos)),
        ("So no sistema (nao na planilha)", len(so_sis)),
    ]
    for label, val in metricas:
        fill = COR_FALTA if label.startswith("FALTANDO") and val > 0 else None
        ws.cell(row=row, column=1, value=label).font = Font(bold=label.startswith("FALTANDO"))
        c2 = ws.cell(row=row, column=2, value=val)
        c2.font = Font(bold=label.startswith("FALTANDO"))
        if fill:
            ws.cell(row=row, column=1).fill = fill
            c2.fill = fill
        row += 1
    row += 1

    # Seção FALTANDO
    if faltando:
        titulo_secao(ws, row, f"FALTANDO NO SISTEMA — {len(faltando)} NFs (importar no sistema)", NCOLS, COR_FALTA)
        row += 1
        cabecalho(ws, row, COLS_FALTA)
        row += 1
        for nf_num in sorted(faltando):
            entrada = next((e for e in entradas if e["numero_nota"] == nf_num), {})
            cel(ws, row, 1, nf_num, COR_FALTA, bold=True, align="center")
            cel(ws, row, 2, entrada.get("condominio", ""), COR_FALTA)
            cel(ws, row, 3, entrada.get("categoria", ""), COR_FALTA)
            cel(ws, row, 4, entrada.get("parcela", ""), COR_FALTA, align="center")
            cel(ws, row, 5, entrada.get("vencimento", ""), COR_FALTA, align="center")
            cel(ws, row, 6, fmt_r(entrada.get("valor")), COR_FALTA, fmt='R$ #,##0.00', align="right")
            cel(ws, row, 7, "Importar no sistema", COR_FALTA)
            row += 1
        row += 1

    # Seção EM AMBOS
    if em_ambos:
        titulo_secao(ws, row, f"PRESENTES EM AMBOS — {len(em_ambos)} NFs", NCOLS, COR_OK)
        row += 1
        cabecalho(ws, row, COLS_OK)
        row += 1
        for nf_num in sorted(em_ambos):
            info = sistema_por_int[nf_num]
            cel(ws, row, 1, nf_num, COR_OK, align="center")
            cel(ws, row, 2, info["condominio"], COR_OK)
            cel(ws, row, 3, info["tipo"], COR_OK, align="center")
            cel(ws, row, 4, info["vencimento"], COR_OK, align="center")
            cel(ws, row, 5, fmt_r(info["valor"]), COR_OK, fmt='R$ #,##0.00', align="right")
            cel(ws, row, 6, info["status"], COR_OK, align="center")
            cel(ws, row, 7, "", COR_OK)
            row += 1
        row += 1

    # Seção SÓ NO SISTEMA
    if so_sis:
        titulo_secao(ws, row, f"SO NO SISTEMA (nao na planilha) — {len(so_sis)} NFs", NCOLS, COR_SODB)
        row += 1
        cabecalho(ws, row, COLS_SIS)
        row += 1
        for nf_num in so_sis:
            info = sistema_por_int.get(nf_num, {})
            cel(ws, row, 1, nf_num, COR_SODB, align="center")
            cel(ws, row, 2, info.get("condominio", ""), COR_SODB)
            cel(ws, row, 3, info.get("tipo", ""), COR_SODB, align="center")
            cel(ws, row, 4, info.get("vencimento", ""), COR_SODB, align="center")
            cel(ws, row, 5, fmt_r(info.get("valor")), COR_SODB, fmt='R$ #,##0.00', align="right")
            cel(ws, row, 6, info.get("status", ""), COR_SODB, align="center")
            cel(ws, row, 7, "", COR_SODB)
            row += 1

    ajustar(ws, LARG)
    ws.freeze_panes = "A7"

# ─────────────────────────────────────────────
# 5. ABA SISTEMA COMPLETO (todas as NFs por mês)
# ─────────────────────────────────────────────
ws_all = wb_out.create_sheet(title="Sistema Completo")
row = 1
ws_all.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
t = ws_all.cell(row=row, column=1, value="Todas as NFs do sistema — notas_fiscais_2026-05-27.xlsx")
t.font = Font(bold=True, size=13)
t.fill = COR_TITULO
t.alignment = Alignment(horizontal="center")
row += 1
ws_all.cell(row=row, column=1, value=f"Total: {len(sistema_lista)} NFs")
row += 2

mes_anterior = None
for entry in sorted(sistema_lista, key=lambda x: (x["mes_venc"] or 99, x["numero_nota"])):
    mes = entry["mes_venc"]
    if mes != mes_anterior:
        mes_anterior = mes
        titulo_secao(ws_all, row, f"Vencimento: {MESES.get(mes, str(mes))} 2026", 7, COR_TITULO)
        row += 1
        cabecalho(ws_all, row, ["Nº NF", "Condominio", "Tipo", "Vencimento", "Valor (R$)", "Status", ""])
        row += 1

    cel(ws_all, row, 1, entry["numero_nota"], align="center")
    cel(ws_all, row, 2, entry["condominio"])
    cel(ws_all, row, 3, entry["tipo"], align="center")
    cel(ws_all, row, 4, entry["vencimento"], align="center")
    cel(ws_all, row, 5, fmt_r(entry["valor"]), fmt='R$ #,##0.00', align="right")
    cel(ws_all, row, 6, entry["status"], align="center")
    cel(ws_all, row, 7, "")
    row += 1

ajustar(ws_all, LARG)
ws_all.freeze_panes = "A5"

# ─────────────────────────────────────────────
# 6. ABA RESUMO (primeira)
# ─────────────────────────────────────────────
ws_res = wb_out.create_sheet(title="Resumo", index=0)
row = 1
ws_res.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
t = ws_res.cell(row=row, column=1, value="Relatorio Comparativo NFs 2026")
t.font = Font(bold=True, size=14)
t.fill = COR_TITULO
t.alignment = Alignment(horizontal="center")
row += 1
ws_res.cell(row=row, column=1, value=f"Gerado em {datetime.now().strftime('%d/%m/%Y %H:%M')}")
row += 2

cabecalho(ws_res, row, ["Mes", "Na Planilha", "No Sistema", "FALTANDO", "Em Ambos", "So no Sistema"])
row += 1

total_plan = total_sis = total_falt = total_ok = total_so = 0
for nome_mes, pl, sis, falt, ok, so in resumo_dados:
    fill = COR_FALTA if falt > 0 else COR_OK
    cel(ws_res, row, 1, nome_mes, fill, bold=True)
    cel(ws_res, row, 2, pl, fill, align="center")
    cel(ws_res, row, 3, sis, fill, align="center")
    cel(ws_res, row, 4, falt, fill, bold=(falt > 0), align="center")
    cel(ws_res, row, 5, ok, fill, align="center")
    cel(ws_res, row, 6, so, fill, align="center")
    row += 1
    total_plan += pl; total_sis += sis; total_falt += falt; total_ok += ok; total_so += so

# Total
for c, v in enumerate([("TOTAL", total_plan, total_sis, total_falt, total_ok, total_so)], 0):
    for i, val in enumerate(v, 1):
        c2 = ws_res.cell(row=row, column=i, value=val)
        c2.font = Font(bold=True)
        c2.border = BORDA

row += 2
if total_falt > 0:
    ws_res.cell(row=row, column=1, value="NFs a importar no sistema:").font = Font(bold=True)
    row += 1
    for mes_num in sorted(planilha_nfs.keys()):
        entradas = planilha_nfs[mes_num]
        nfs_plan = set(e["numero_nota"] for e in entradas)
        falt = sorted(n for n in nfs_plan if n not in sistema_por_int)
        if falt:
            c2 = ws_res.cell(row=row, column=1,
                value=f"{MESES.get(mes_num)}: NFs {min(falt)} a {max(falt)} ({len(falt)} notas)")
            c2.fill = COR_FALTA
            c2.font = Font(bold=True)
            row += 1

ajustar(ws_res, [16, 14, 14, 14, 12, 16])

# ─────────────────────────────────────────────
# 7. SALVAR
# ─────────────────────────────────────────────
out_path = BASE / "RELATORIO_NF_2026_v2.xlsx"
wb_out.save(out_path)
print(f"\nOK Relatorio salvo: {out_path}")
print(f"   Abas: Resumo | {' | '.join(MESES.get(m) for m in sorted(planilha_nfs.keys()))} | Sistema Completo")
print(f"   Total FALTANDO: {total_falt} NFs")
