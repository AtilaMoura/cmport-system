import openpyxl, datetime, unicodedata, sys

COND_IDS = {
    'Assumpita Sica': 378, 'Alvorada': 225, 'Bahamas': 596, 'Bonaire': 663,
    'Cap Martinique': 414, 'Cezario Motta': 670, 'Cintia': 650, 'Costa Brava': 394,
    'Cube Vila Ipojuca': 714, 'Ipiranga': 176, 'Lucerna': 651, 'Odete': 689,
    'Olivais': 661, 'Paco de Hygienopolis': 675, 'Penthouse': 671,
    'Piazza Fontana': 678, 'Raposo Tavares': 672, 'Sbios': 451,
    'Sintonia Perdizes': 534, 'The Crystal Houser': 505, 'Ville Courchevel': 244,
    'Angra dos Reis': 648, 'Concor': 685, 'Maison Saint Etienne': 674,
    'Park': 673, 'Vermont': 652, 'Via Del Corso': 657,
    'Helbor Lof Evolution': 679, 'Helbor Loft': 679, 'Helbor Loft Evolution': 679,
    'Ibirapuera Park': 684, 'Jardim Buenos Aires': 653, 'Park Avenue': 676,
    'Reynolds': 654, 'Fortezza Di Ferrara': 742, 'Jose Antonio Alpiovezza': 686,
    'Jussara': 668, 'Jussara - Mes de Dezembro': 668, 'Le Monde': 665,
    'Macunaima': 125, 'Margaux': 152, 'Mont Blanc': 688, 'Verbena': 175,
    'Araucarias': 582, 'Cullinan': 154, 'Cullian': 154, 'Olga': 736,
    'Mariangela Texeira': 555, 'Villar Paraiso': 728, 'Green Gold': 384,
    'San Mariano': 139, 'Lorrenzetti': 233, 'Sao Vicente': 759,
    'Zueno (Tapajos)': 484, 'Bambino': 578, 'Five Stars': 365,
    'Castel Mantova': 149, 'California Park': 529,
    'Pracas Sao Paulo': 219, 'Fabiola': 454, 'Porto Alegre': 501,
    'Grevillle Flamboyate': 733, 'The Cristal': 505,
    'Maison Saint Etiene': 674, 'Pacos Hygeanopolis': 675,
    'Dulce': 730, 'BBARE 02/06': 602, 'BBARE 03/06': 602, 'BBARE todos': 602,
    'Edgar': 148, 'Mario Cantoni': 333, 'Alessandra': 608,
    'Helbor Lof Evolution (Rodrigo ap. 605)': 679, 'Rafaela': 469,
}
PENDING_NAMES = {'Eraseg', 'Durval', 'Adelson', 'Ludmila', 'Luis', 'Chistopher'}

def norm(s):
    nfkd = unicodedata.normalize('NFKD', str(s).strip())
    return ''.join(c for c in nfkd if not unicodedata.combining(c))

def map_cond(raw):
    key = norm(raw).strip()
    for k, v in COND_IDS.items():
        if norm(k).lower() == key.lower():
            return v
    for k, v in COND_IDS.items():
        nk = norm(k).lower()
        if nk in key.lower() or key.lower() in nk:
            return v
    return None

def fmt(d):
    if isinstance(d, datetime.datetime):
        return d.strftime('%Y-%m-%d')
    return str(d) if d else None

def esc(s):
    return str(s).replace("'", "''").replace("\\", "\\\\")

wb = openpyxl.load_workbook('Entradas Fluxo Janeiro.xlsx', data_only=True)
ws = wb.active
SKIP = {1, 2, 3, 50, 51, 52, 53, 124}
recibo_counter = 0

lines = [
    'SET NAMES utf8mb4;',
    'SET autocommit=0;',
    'START TRANSACTION;',
    '',
]
inserted = 0
pending_rows = []  # (r, raw_cond, motivo)

for r in range(4, ws.max_row + 1):
    if r in SKIP:
        continue
    v = [ws.cell(r, c).value for c in range(1, 13)]
    if not any(vv is not None for vv in v):
        continue

    raw_cond = str(v[2]).strip() if v[2] else ''
    categoria = str(v[3]).strip() if v[3] else ''
    nf_raw    = v[4]
    parcela   = str(v[5]).strip() if v[5] else '01/01'
    dt_pagto  = v[6]
    dt_vencto = v[7]
    valor     = v[8]

    norm_cond = norm(raw_cond)
    is_recibo = str(nf_raw).strip().lower() == 'recibo'
    is_pending = any(norm(p).lower() in norm_cond.lower() for p in PENDING_NAMES)

    if is_pending:
        pending_rows.append((r, raw_cond, 'Condominio nao encontrado no sistema'))
        continue

    cond_id = map_cond(raw_cond)
    if cond_id is None:
        pending_rows.append((r, raw_cond, 'Sem mapeamento de condominio'))
        continue

    if is_recibo:
        recibo_counter += 1
        nf_num = 'REC-2026-{:03d}'.format(recibo_counter)
    else:
        nf_num = str(nf_raw).strip()

    try:
        p1, p2 = parcela.split('/')
        parc_atual = int(p1)
        parc_total = int(p2)
    except Exception:
        parc_atual = 1
        parc_total = 1

    tipo_nf = 'MANUTENCAO' if categoria == 'Contrato' else 'ASSISTENCIA'
    tipo_ma = 'manutencao' if categoria == 'Contrato' else 'assistencia'
    d_pag = fmt(dt_pagto)
    d_vec = fmt(dt_vencto)
    val = round(float(valor), 2) if valor else 0.0

    lines.append('-- Row {}: {} | NF: {} | cond_id: {}'.format(r, raw_cond, nf_num, cond_id))
    lines.append(
        "INSERT INTO notas_fiscais\n"
        "  (numero_nota, tipo, status, parcelas, valor, data_vencimento, data_pagamento, condominio_id, xml_original, alerta_impostos)\n"
        "  VALUES ('{}', '{}', 'AUTORIZADA', 1, {}, '{}', '{}', {}, 'ENTRADA_MANUAL', 0)\n"
        "  ON DUPLICATE KEY UPDATE data_pagamento=VALUES(data_pagamento), id=LAST_INSERT_ID(id);".format(
            esc(nf_num), tipo_nf, val, d_vec, d_pag, cond_id
        )
    )
    lines.append('SET @nf_id = LAST_INSERT_ID();')
    lines.append(
        "INSERT INTO manutencoes_assistencias\n"
        "  (condominio_id, nota_fiscal_id, tipo, data_servico, criado_em, atualizado_em)\n"
        "  VALUES ({}, @nf_id, '{}', '{}', NOW(), NOW());".format(cond_id, tipo_ma, d_pag)
    )
    lines.append(
        "INSERT INTO boletos\n"
        "  (nota_fiscal_id, numero_parcela, total_parcelas, valor_nominal, valor_total_recebido,\n"
        "   data_emissao, data_vencimento, data_pagamento, tipo_cobranca, situacao, forma_pagamento,\n"
        "   valor_juros, valor_multa)\n"
        "  VALUES (@nf_id, 1, 1, {}, {}, '{}', '{}', '{}', 'SIMPLES', 'PAGO', 'TRANSFERENCIA', 0.0, 0.0);".format(
            val, val, d_pag, d_vec, d_pag
        )
    )
    lines.append('')
    inserted += 1

lines.append('COMMIT;')
lines.append('-- Total inseridos: {}, pendentes: {}'.format(inserted, len(pending_rows)))

sql = '\n'.join(lines)
with open('insercao_janeiro_2026.sql', 'w', encoding='utf-8') as f:
    f.write(sql)

# Salvar pendentes para atualizar XLSX depois
with open('pendentes_janeiro.txt', 'w', encoding='utf-8') as f:
    for pr in pending_rows:
        f.write('{}\t{}\t{}\n'.format(*pr))

print('SQL gerado: {} registros, {} pendentes'.format(inserted, len(pending_rows)))
print('Pendentes:')
for pr in pending_rows:
    print('  Row {}: {} — {}'.format(*pr))
