# -*- coding: utf-8 -*-
"""
Passo 1 do PLANO_RECONCILIACAO_EXTRATOS_AGOSTO.md — parser unico dos 3 extratos
bancarios de agosto/2026. So LEITURA, nao toca banco nenhum.

Entradas:
  docs-e-planilhas/Extrato Inter.csv       -> banco_id 2  (Inter CMPORT, conta 308310110)
  docs-e-planilhas/Extrato Inter tec.csv   -> banco_id 4  (Inter TEC,    conta 524203806)
  docs-e-planilhas/Extrato Itau.xlsx       -> banco_id 1  (Itau CMPORT,  ag 8135 cc 17278-4)

Saida:
  - normaliza tudo em {banco_id, conta, data, descricao, valor(+/-), tipo, agregado}
  - classifica: ENTRADA | SAIDA | TRANSFERENCIA_ENTRE_CONTAS | TARIFA | RENDIMENTO | DEBITO_CARTAO
  - grava fluxo-financeiro/extratos_agosto_normalizado.json
  - imprime resumo por conta (Σ entrada/saida/transf, conferencia de saldo)

Uso: cd backend && ./venv/Scripts/python.exe ../fluxo-financeiro/ler_extratos_agosto.py
"""
import io
import json
import os
import re
import sys
from datetime import date

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
DOCS = os.path.join(BASE, "docs-e-planilhas")
OUT = os.path.join(BASE, "fluxo-financeiro", "extratos_agosto_normalizado.json")

# conta do extrato -> (banco_id, rotulo)
CONTAS = {
    "308310110": (2, "Inter CMPORT"),
    "524203806": (4, "Inter TEC"),
    "itau":      (1, "Itau CMPORT"),
}
# numeros de conta que identificam transferencia interna entre as 2 contas Inter
CONTAS_INTERNAS = {"308310110", "524203806"}
# CNPJs do grupo (CMPORT / CMPORT TEC / "C&M PORT" Itau)
CNPJS_GRUPO = {"22761557000188", "65756913000188"}


def brl(s):
    """'1.882,24' / '-543' -> float"""
    s = str(s).strip().replace(".", "").replace(",", ".")
    return round(float(s), 2)


def fix_cp1252(s):
    """Conserta mojibake cp1252 do Itau (mesmo criterio do corrigir_encoding_historico.py)."""
    if not s:
        return s
    try:
        cand = s.encode("cp1252").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return s
    sus = "ÃÂ¢â‚¬œ�"
    return cand if sum(c in sus for c in cand) < sum(c in sus for c in s) else s


def classificar(desc, valor, agregado, cnpj_contraparte=""):
    d = desc.lower()
    # transferencia entre as 2 contas Inter (Pix com o numero da outra conta)
    if any(c in desc for c in CONTAS_INTERNAS):
        return "TRANSFERENCIA_ENTRE_CONTAS"
    # Itau: pix enviado/recebido cuja contraparte e' CNPJ do proprio grupo = transf interna
    so_num = re.sub(r"\D", "", cnpj_contraparte)
    if "pix" in d and so_num and any(c in so_num for c in CNPJS_GRUPO):
        return "TRANSFERENCIA_ENTRE_CONTAS"
    if any(k in d for k in ("tar plano", "tar ", " iof", "iof\b", "juros limite",
                            "tarifa", "cesta ", "manutencao de conta")):
        return "TARIFA"
    if d.strip().startswith("iof"):
        return "TARIFA"
    if "rendiment" in d or "rend pago aplic" in d or "rend. pago" in d:
        return "RENDIMENTO"
    if "compra no debito" in d:
        return "DEBITO_CARTAO"
    return "ENTRADA" if valor > 0 else "SAIDA"


def ler_inter(path, conta):
    """
    O Inter da 'Saldo:' no cabecalho = saldo AO VIVO (dia da geracao do extrato,
    ja com movimento de setembro) — nao serve pra fechar agosto. O saldo real de
    fim de periodo e' a coluna 'Saldo' (running balance) do ULTIMO lancamento.
    saldo inicial = (saldo running do 1o lancamento) - (valor do 1o lancamento).
    """
    banco_id, rotulo = CONTAS[conta]
    itens, saldo_ini, saldo_fim, saldo_vivo = [], None, None, None
    with open(path, encoding="utf-8") as f:
        linhas = f.read().splitlines()
    primeiro = True
    for ln in linhas:
        p = ln.split(";")
        if len(p) < 4:
            continue
        if p[0].strip() == "Saldo:":
            saldo_vivo = brl(p[1])
            continue
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})$", p[0].strip())
        if not m:
            continue
        d, mo, y = map(int, m.groups())
        desc = p[1].strip().strip('"')
        valor = brl(p[2])
        running = brl(p[3])
        if primeiro:
            saldo_ini = round(running - valor, 2)
            primeiro = False
        saldo_fim = running  # sobrescreve ate o ultimo
        itens.append({
            "banco_id": banco_id, "conta": conta, "rotulo": rotulo,
            "data": date(y, mo, d).isoformat(),
            "descricao": desc, "valor": valor, "agregado": False,
            "tipo": classificar(desc, valor, False),
        })
    return itens, saldo_ini, saldo_fim, saldo_vivo


def ler_itau(path):
    import openpyxl
    banco_id, rotulo = CONTAS["itau"]
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    itens, saldo_ini, saldo_fim = [], None, None
    for row in ws.iter_rows(values_only=True):
        if not row or not row[0]:
            continue
        c0 = str(row[0]).strip()
        m = re.match(r"(\d{2})/(\d{2})/(\d{4})$", c0)
        if not m:
            continue
        lanc = fix_cp1252(str(row[1] or "").strip())
        razao = fix_cp1252(str(row[2] or "").strip())
        cnpj_cp = str(row[3] or "").strip()
        valor_raw, saldo_raw = row[4], row[5]
        d, mo, y = map(int, m.groups())
        dt = date(y, mo, d)
        if lanc.upper() == "SALDO ANTERIOR":
            saldo_ini = round(float(saldo_raw), 2) if saldo_raw is not None else None
            continue
        if "SALDO TOTAL DISPON" in lanc.upper():
            if saldo_raw is not None:
                saldo_fim = round(float(saldo_raw), 2)
            continue
        if valor_raw is None:
            continue
        valor = round(float(valor_raw), 2)
        agregado = "BOLETOS RECEBIDOS" in lanc.upper()
        desc = f"{lanc} — {razao}".strip(" —") if razao else lanc
        itens.append({
            "banco_id": banco_id, "conta": "itau", "rotulo": rotulo,
            "data": dt.isoformat(),
            "descricao": desc, "razao_social": razao, "cnpj_contraparte": cnpj_cp,
            "valor": valor, "agregado": agregado,
            "tipo": classificar(desc, valor, agregado, cnpj_cp),
        })
    return itens, saldo_ini, saldo_fim


def resumo(nome, itens, saldo_ini, saldo_fim):
    por_tipo = {}
    for it in itens:
        por_tipo.setdefault(it["tipo"], [0, 0.0])
        por_tipo[it["tipo"]][0] += 1
        por_tipo[it["tipo"]][1] += it["valor"]
    soma = round(sum(it["valor"] for it in itens), 2)
    print(f"\n{'='*70}\n{nome}   ({len(itens)} lançamentos)")
    print(f"{'='*70}")
    for tp, (n, v) in sorted(por_tipo.items(), key=lambda x: -x[1][1]):
        print(f"  {tp:28} {n:3}   R$ {v:>13,.2f}")
    print(f"  {'-'*28} {'':3}   {'-'*16}")
    print(f"  {'Σ movimento do mês':28} {len(itens):3}   R$ {soma:>13,.2f}")
    if saldo_ini is not None and saldo_fim is not None:
        calc = round(saldo_ini + soma, 2)
        ok = "✅" if abs(calc - saldo_fim) < 0.02 else "❌"
        print(f"  saldo inicial  R$ {saldo_ini:>13,.2f}")
        print(f"  saldo final    R$ {saldo_fim:>13,.2f}  (extrato)")
        print(f"  saldo ini + Σ  R$ {calc:>13,.2f}  {ok} "
              f"{'' if ok=='✅' else f'dif R$ {calc-saldo_fim:,.2f}'}")


def main():
    tudo = []
    inter_c, si_c, sf_c, sv_c = ler_inter(os.path.join(DOCS, "Extrato Inter.csv"), "308310110")
    inter_t, si_t, sf_t, sv_t = ler_inter(os.path.join(DOCS, "Extrato Inter tec.csv"), "524203806")
    itau, si_i, sf_i = ler_itau(os.path.join(DOCS, "Extrato Itau.xlsx"))
    print(f"(Inter 'Saldo:' ao vivo — CMPORT {sv_c} · TEC {sv_t} — não usado p/ fechar agosto)")

    resumo("INTER CMPORT (banco 2, conta 308310110)", inter_c, si_c, sf_c)
    resumo("INTER TEC (banco 4, conta 524203806)", inter_t, si_t, sf_t)
    resumo("ITAU CMPORT (banco 1, ag 8135 cc 17278-4)", itau, si_i, sf_i)

    tudo = inter_c + inter_t + itau
    with open(OUT, "w", encoding="utf-8") as f:
        json.dump({
            "gerado_em": "agosto/2026",
            "contas": {
                "308310110": {"banco_id": 2, "saldo_inicial": si_c, "saldo_final": sf_c},
                "524203806": {"banco_id": 4, "saldo_inicial": si_t, "saldo_final": sf_t},
                "itau":      {"banco_id": 1, "saldo_inicial": si_i, "saldo_final": sf_i},
            },
            "lancamentos": tudo,
        }, f, ensure_ascii=False, indent=1)
    print(f"\n→ {OUT}  ({len(tudo)} lançamentos)")

    # amostra das transferencias internas (pro Passo 3)
    tr = [it for it in tudo if it["tipo"] == "TRANSFERENCIA_ENTRE_CONTAS"]
    print(f"\nTRANSFERÊNCIAS ENTRE CONTAS detectadas: {len(tr)}")
    for it in sorted(tr, key=lambda x: (x["conta"], x["data"])):
        print(f"  {it['data']}  {it['rotulo']:13} R$ {it['valor']:>11,.2f}  {it['descricao'][:52]}")


if __name__ == "__main__":
    main()
