# -*- coding: utf-8 -*-
"""
Extrai a planilha "Controle de Funcionarios - 2026.xlsx" (aba Pagamentos - 2026)
para JSON estruturado + CSV plano.

A aba tem layout em blocos: um bloco por mes (JANEIRO..DEZEMBRO), e dentro de
cada mes os funcionarios ficam lado a lado em 3 faixas de colunas
(1-4, 6-9, 11-14). Cada funcionario tem cabecalho (NOME / ENTRADA / SALARIO),
uma tabela DATA|VALES|FALTAS|ASSUNTO e uma linha de TOTAL (liquido pago no mes).

Saidas (em fluxo-financeiro/):
  - controle_funcionarios_2026.json   (aninhado: mes -> funcionario -> itens)
  - controle_funcionarios_2026_itens.csv   (1 linha por lancamento)
  - controle_funcionarios_2026_totais.csv  (1 linha por funcionario/mes: total liquido)

Tambem le as abas Ferias / Folgas / CA e joga no JSON (campos rh_*).
"""
import json
import csv
import re
import unicodedata
from pathlib import Path

import openpyxl

BASE = Path(__file__).resolve().parent
XLSX = BASE.parent / "docs-e-planilhas" / "Controle de Funcionarios - 2026.xlsx"

MESES = {
    "JANEIRO": 1, "FEVEREIRO": 2, "MARCO": 3, "MARÇO": 3, "ABRIL": 4,
    "MAIO": 5, "JUNHO": 6, "JULHO": 7, "AGOSTO": 8, "SETEMBRO": 9,
    "OUTUBRO": 10, "NOVEMBRO": 11, "DEZEMBRO": 12,
}

# faixas de colunas dos 3 funcionarios lado a lado (base = coluna da DATA)
GRUPOS = [1, 6, 11]

# ---------------------------------------------------------------------------

def _txt(v):
    if v is None:
        return ""
    return str(v).strip()


def _sem_acento(s):
    return "".join(
        c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn"
    ).upper()


def _num(v):
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return round(float(v), 2)
    s = str(v).strip().replace(".", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None


DATE_RE = re.compile(r"^(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{2,6})$")


def _data(v):
    """Retorna ISO 'YYYY-MM-DD' ou None. Tolera '12.03.20260' (typo da planilha)."""
    if v is None:
        return None
    if hasattr(v, "isoformat"):
        try:
            return v.date().isoformat()
        except Exception:
            return v.isoformat()[:10]
    s = str(v).strip()
    m = DATE_RE.match(s)
    if not m:
        return None
    d, mth, y = m.groups()
    y = y[:4] if len(y) > 4 else y            # '20260' -> '2026'
    if len(y) == 2:
        y = "20" + y
    try:
        return f"{int(y):04d}-{int(mth):02d}-{int(d):02d}"
    except ValueError:
        return None


# nome canonico dos 7 funcionarios --------------------------------------------
CANON = {
    "LUIS ANTONIO MELGAREJO NEVES": "Luis Antonio Melgarejo Neves",
    "WELLIGTON LUCAS MENESES RODRIGUES": "Welligton Lucas Meneses Rodrigues",
    "WELLIGTON LUCAS MENEZES RODRIGUES": "Welligton Lucas Meneses Rodrigues",
    "ANDRE MOREIRA ROSA": "Andre Moreira Rosa",
    "FABIANA PEDRETTI": "Fabiana Pedretti Moreira Rosa",
    "FABIANA PEDRETTI MOREIRA ROSA": "Fabiana Pedretti Moreira Rosa",
    "PEDRO HENRIQUE DA SILVA": "Pedro Henrique da Silva",
    "GABRIEL MOREIRA PEDRETTI": "Gabriel Moreira Pedretti",
    "ALMIRA MOPREIRA ROSA SALOMAO": "Almira Moreira Rosa Salomao",
    "ALMIRA MOREIRA ROSA SALOMAO": "Almira Moreira Rosa Salomao",
}


def canonizar(nome_raw):
    key = _sem_acento(nome_raw).strip()
    key = re.sub(r"\s+", " ", key)
    return CANON.get(key, nome_raw.strip())


def eh_nome(c0):
    if not c0:
        return False
    up = _sem_acento(c0)
    if up.startswith(("ENTRADA", "SALARIO", "SALARIA", "PRE ", "PRE-", "PRE- ", "DATA",
                      "VALES DE FUNCIONARIOS", "PROGRAMACAO", "FOLGAS", "CA DOS")):
        return False
    # nome tem pelo menos 2 palavras e nenhuma e numero
    partes = c0.split()
    return len(partes) >= 2 and not any(p.replace(".", "").isdigit() for p in partes)


# classificacao do lancamento a partir do texto do ASSUNTO --------------------
def classificar(assunto, valor):
    a = _sem_acento(assunto)
    neg = (valor is not None and valor < 0)
    regras = [
        ("SALARIO", ("SALARIO", "DIAS TRABALHADOS", "PRO LABORE", "PRO-LABORE")),
        ("ADIANTAMENTO", ("ADIANTAMENTO",)),
        ("VALE_REFEICAO", ("VALE REFEICAO", "REFEICAO")),
        ("VALE_ALIMENTACAO", ("VALE ALIMENTACAO", "ALIMENTACAO")),
        ("VALE_TRANSPORTE", ("VALE TRANSPORTE", "TRANSPORTE", " VT ", "VT ")),
        ("PLANTAO", ("PLANTAO", "PALNTAO", "PANTAO", "PLANTA")),
        ("HORA_EXTRA", ("HORA EXTRA", "HORAS EXTRA", "EXTRAS", "HORAS  EXTRA")),
        ("REFLEXO_DSR", ("REFLEXO", "RERLEXO", "D.S.R", "DSR")),
        ("BONIFICACAO", ("BONIFICA", "CHAMADO", "CHAAMADO", "CAHAMADO")),
        ("INSS", ("INSS",)),
        ("IRRF", ("IMPOSTO DE RENDA", "IRRF", "I.R.R.F")),
        ("CONTRIB_ASSISTENCIAL", ("CONTRIBUICAO ASSISTENCIAL", "CONTRIBUICAO ASSSITENCIAL",
                                  "CONTRIBUICAO ASSITENCIAL")),
        ("DESCONTO_13", ("MEDIA  HORA 13", "MEDIA HORA 13", "DIFERENCA VALOR 13", "13")),
        ("MONITOR", ("MONITOR",)),
        ("MULTA", ("MULTA",)),
        ("HABILITACAO", ("HABILITACAO",)),
        ("REEMBOLSO", ("REEMBOLSO", "REMBOLSO")),
        ("DESCONTO_VR_VT", ("VR NAO", "VT NAO", "NAO ULTILIZADO", "NAO ULTIZADO")),
    ]
    for tipo, chaves in regras:
        for k in chaves:
            if k.strip() in a:
                return tipo
    return "DESCONTO" if neg else "OUTRO"


# ---------------------------------------------------------------------------

def extrair_pagamentos(ws):
    """Retorna dict: {mes_num: {"mes_nome":..., "funcionarios":[...]}}"""
    # 1) achar as linhas de inicio de cada mes
    marcos = []  # (row, mes_num, mes_nome)
    for r in range(1, ws.max_row + 1):
        c0 = _txt(ws.cell(r, 1).value)
        m = re.match(r"VALES DE FUNCIONARIOS\s*-\s*([A-ZÇÃÁÉÍ]+)\s*-\s*2026", _sem_acento(c0) and c0 or "")
        if c0.upper().startswith("VALES DE FUNCIONARIOS"):
            mp = re.search(r"-\s*([A-Za-zÇÃÁÉÍç]+)\s*-\s*2026", c0)
            nome = mp.group(1).upper() if mp else ""
            num = MESES.get(nome) or MESES.get(_sem_acento(nome))
            if num:
                marcos.append((r, num, nome))
    marcos.append((ws.max_row + 1, None, None))

    resultado = {}
    for i in range(len(marcos) - 1):
        ini, mes_num, mes_nome = marcos[i]
        fim = marcos[i + 1][0]
        funcs = []
        for base in GRUPOS:
            funcs.extend(_extrair_grupo(ws, ini, fim, base))
        resultado[mes_num] = {"mes_nome": mes_nome.capitalize(), "funcionarios": funcs}
    return resultado


def _extrair_grupo(ws, ini, fim, base):
    """Extrai todos os registros de funcionario numa faixa de colunas [base..base+3]."""
    out = []
    r = ini
    atual = None
    while r < fim:
        c0 = _txt(ws.cell(r, base).value)
        c1 = ws.cell(r, base + 1).value
        c2 = ws.cell(r, base + 2).value
        c3 = _txt(ws.cell(r, base + 3).value)

        if eh_nome(c0):
            if atual:
                out.append(atual)
            atual = {
                "nome": canonizar(c0),
                "nome_planilha": c0,
                "entrada_raw": "",
                "salario_raw": "",
                "itens": [],
                "total_liquido": None,
            }
        elif atual is not None:
            up = _sem_acento(c0)
            if up.startswith("ENTRADA"):
                extra = _txt(c1) or _txt(ws.cell(r, base + 1).value)
                atual["entrada_raw"] = (c0 + (" " + extra if extra else "")).strip()
            elif up.startswith(("SALARIO", "SALARIA", "PRE ", "PRE-", "PRE")):
                atual["salario_raw"] = c0
            elif up == "DATA":
                pass  # cabecalho da tabela
            else:
                data_iso = _data(ws.cell(r, base).value)
                valor = _num(c1)
                if data_iso and valor is not None:
                    atual["itens"].append({
                        "data": data_iso,
                        "valor": valor,
                        "falta": _txt(c2) or None,
                        "assunto": c3,
                        "tipo": classificar(c3, valor),
                    })
                elif not c0 and valor is not None and not c3:
                    # linha de TOTAL (so a coluna de valor preenchida)
                    atual["total_liquido"] = valor
        r += 1
    if atual:
        out.append(atual)
    return out


def _linhas_do_bloco(ws, r_ini, base):
    """Rows de um bloco de funcionario numa faixa de colunas, ate o proximo nome."""
    linhas = []
    rr = r_ini
    while rr <= ws.max_row:
        x0 = _txt(ws.cell(rr, base).value)
        up = _sem_acento(x0)
        if rr != r_ini and (eh_nome(x0) or up.startswith(
            ("PROGRAMACAO", "FOLGAS DOS", "CA DOS", "VALES DE")
        )):
            break
        cells = [ws.cell(rr, base + k).value for k in range(4)]
        if any(_txt(c) for c in cells) and not up.startswith(
            ("ENTRADA", "SALARIO", "SALARIA", "PRE ", "PRE-", "PRE")
        ):
            linhas.append([_txt(c) for c in cells])
        rr += 1
    return linhas


def extrair_ferias(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        for base in (1, 6, 11):
            c0 = _txt(ws.cell(r, base).value)
            if not eh_nome(c0):
                continue
            linhas = _linhas_do_bloco(ws, r + 1, base)
            data_limite, dias, notas = None, [], []
            for cells in linhas:
                a, b = cells[0], cells[1]
                ai = _data(a)
                if ai and not data_limite:
                    data_limite = ai
                elif ai:
                    notas.append({"data": ai, "detalhe": cells[3] or b})
                elif a and a.upper() not in ("DATA LIMITE", "DATA", "VALES", "DIAS") and not _num(a):
                    notas.append({"detalhe": a})
                if b and b != "0" and "DIA" in b.upper():
                    dias.append(b)
            out[canonizar(c0)] = {
                "data_limite_periodo": data_limite,
                "dias_programados": dias,
                "notas": notas,
            }
    return out


def extrair_folgas(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        for base in (1, 6, 11):
            c0 = _txt(ws.cell(r, base).value)
            if not eh_nome(c0):
                continue
            linhas = _linhas_do_bloco(ws, r + 1, base)
            registros = []
            for cells in linhas:
                data, _v, _f, assunto = cells
                if cells[0].upper() in ("DATA", "VALES", "FALTAS", "ASSUNTO"):
                    continue
                txt = " ".join(x for x in (data, assunto) if x and x != "0")
                if txt.strip():
                    registros.append({"data": _data(data), "detalhe": assunto or data})
            out[canonizar(c0)] = registros
    return out


def extrair_ca(ws):
    out = {}
    for r in range(1, ws.max_row + 1):
        for base in (1, 6, 11):
            c0 = _txt(ws.cell(r, base).value)
            if not eh_nome(c0):
                continue
            linhas = _linhas_do_bloco(ws, r + 1, base)
            epis = []
            for cells in linhas:
                data, qtd, tipo, ca_val = cells
                if not _data(data):
                    continue
                ca, validade = ca_val, None
                m = re.match(r"\s*(.+?)\s*-\s*(\d{1,2}/\d{1,2}/\d{2,4})\s*$", ca_val)
                if m:
                    ca, validade = m.group(1).strip(), _data(m.group(2))
                epis.append({
                    "data_entrega": _data(data),
                    "quantidade": _num(qtd) or qtd,
                    "tipo": tipo,
                    "ca": ca,
                    "validade": validade,
                })
            out[canonizar(c0)] = epis
    return out


# ---------------------------------------------------------------------------

def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    pag = extrair_pagamentos(wb["Pagamentos - 2026"])

    rh = {
        "ferias": extrair_ferias(wb["Ferias - 2026"]) if "Ferias - 2026" in wb.sheetnames else {},
        "folgas": extrair_folgas(wb["Folgas Feriados - 2025"]) if "Folgas Feriados - 2025" in wb.sheetnames else {},
        "ca_equipamentos": extrair_ca(wb["CA - Equipamentos"]) if "CA - Equipamentos" in wb.sheetnames else {},
    }

    # ---- JSON aninhado ----
    doc = {
        "origem": XLSX.name,
        "gerado_por": "extrair_controle_funcionarios.py",
        "meses": pag,
        "rh": rh,
    }
    out_json = BASE / "controle_funcionarios_2026.json"
    out_json.write_text(json.dumps(doc, ensure_ascii=False, indent=2), encoding="utf-8")

    # ---- CSV plano de itens ----
    itens_csv = BASE / "controle_funcionarios_2026_itens.csv"
    with itens_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["mes", "mes_nome", "funcionario", "data_pagamento", "valor",
                    "tipo", "falta", "assunto", "salario_raw", "entrada_raw"])
        for mnum in sorted(k for k in pag if k):
            bloco = pag[mnum]
            for fn in bloco["funcionarios"]:
                for it in fn["itens"]:
                    w.writerow([mnum, bloco["mes_nome"], fn["nome"], it["data"],
                               f'{it["valor"]:.2f}'.replace(".", ","), it["tipo"],
                               it["falta"] or "", it["assunto"],
                               fn["salario_raw"], fn["entrada_raw"]])

    # ---- CSV de totais por funcionario/mes ----
    tot_csv = BASE / "controle_funcionarios_2026_totais.csv"
    with tot_csv.open("w", newline="", encoding="utf-8-sig") as f:
        w = csv.writer(f, delimiter=";")
        w.writerow(["mes", "mes_nome", "funcionario", "total_liquido",
                    "soma_itens", "diferenca", "qtd_itens"])
        for mnum in sorted(k for k in pag if k):
            bloco = pag[mnum]
            for fn in bloco["funcionarios"]:
                soma = round(sum(i["valor"] for i in fn["itens"]), 2)
                tl = fn["total_liquido"]
                dif = round((tl - soma), 2) if tl is not None else None
                w.writerow([mnum, bloco["mes_nome"], fn["nome"],
                            "" if tl is None else f"{tl:.2f}".replace(".", ","),
                            f"{soma:.2f}".replace(".", ","),
                            "" if dif is None else f"{dif:.2f}".replace(".", ","),
                            len(fn["itens"])])

    # ---- resumo no stdout ----
    print(f"OK -> {out_json.name}")
    print(f"OK -> {itens_csv.name}")
    print(f"OK -> {tot_csv.name}")
    print()
    tot_itens = 0
    tot_geral = 0.0
    print(f"{'mes':>3}  {'funcionario':<32} {'itens':>5} {'soma_itens':>12} {'total_planilha':>14} {'dif':>9}")
    for mnum in sorted(k for k in pag if k):
        bloco = pag[mnum]
        for fn in bloco["funcionarios"]:
            soma = round(sum(i["valor"] for i in fn["itens"]), 2)
            tl = fn["total_liquido"]
            dif = "" if tl is None else f"{tl - soma:+.2f}"
            tot_itens += len(fn["itens"])
            if tl:
                tot_geral += tl
            print(f"{mnum:>3}  {fn['nome']:<32} {len(fn['itens']):>5} "
                  f"{soma:>12.2f} {('' if tl is None else f'{tl:.2f}'):>14} {dif:>9}")
    print()
    print(f"total de itens extraidos: {tot_itens}")
    print(f"soma dos totais liquidos (planilha): R$ {tot_geral:,.2f}")

    # tipos
    from collections import Counter
    tc = Counter()
    for mnum in pag:
        if not mnum:
            continue
        for fn in pag[mnum]["funcionarios"]:
            for it in fn["itens"]:
                tc[it["tipo"]] += 1
    print("\nlancamentos por tipo:")
    for t, n in tc.most_common():
        print(f"  {t:<22} {n}")


if __name__ == "__main__":
    main()
