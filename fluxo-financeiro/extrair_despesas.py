"""Extrai os blocos 'DESPESAS ESCRITORIO' das planilhas de fluxo financeiro
(CMPORT e TEC), de Janeiro ate o mes mais recente disponivel, e salva um
JSON bruto (uma linha = um lancamento, sem nenhum agrupamento ainda)."""
import json
import re
import openpyxl
from datetime import datetime

ARQUIVOS = {
    "CMPORT": "docs-e-planilhas/FLUXO FINANCEIRO - 2026.xlsx",
    "TEC": "docs-e-planilhas/FLUXO FINANCEIRO CMPORT TEC - 2026.xlsx",
}

COL_ID = 0
COL_COD = 1
COL_DESC = 2
COL_CATEGORIA = 3
COL_NF = 4
COL_PARCELA = 5
COL_PAGTO = 6
COL_VENCTO = 7
COL_PAGOS = 8
COL_VALOR = 10


def serializa(v):
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    return v


def extrair(fname, cnpj_label):
    wb = openpyxl.load_workbook(fname, read_only=True, data_only=True)
    ws = wb["Entradas e SAIDAS - 2026"]
    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    resultado = []
    dentro_bloco = False
    mes_atual = None

    for i, row in enumerate(rows):
        v0 = row[0] if len(row) > 0 else None
        if isinstance(v0, str) and "DESPESAS ESCRITORIO" in v0.upper().replace("�", "O"):
            dentro_bloco = True
            m = re.search(r"M[ÊEÊ�]S\s+([A-Z�]+)\s*-\s*(\d{4})", v0.upper())
            mes_atual = v0
            continue
        if isinstance(v0, str) and v0.upper().startswith("FORNECEDORES"):
            dentro_bloco = False
            continue

        if not dentro_bloco:
            continue

        desc = row[COL_DESC] if len(row) > COL_DESC else None
        categoria = row[COL_CATEGORIA] if len(row) > COL_CATEGORIA else None
        if categoria is None:
            continue
        cat_norm = str(categoria).strip().lower()
        if cat_norm in ("categoria",):  # linha de cabecalho repetida
            continue
        if "escritorio" not in cat_norm:
            continue
        if desc is None:
            continue

        item = {
            "linha_planilha": i + 1,
            "cnpj": cnpj_label,
            "bloco_mes": mes_atual,
            "id_interno": row[COL_ID] if len(row) > COL_ID else None,
            "codigo": row[COL_COD] if len(row) > COL_COD else None,
            "descricao": str(desc).strip(),
            "categoria": str(categoria).strip(),
            "nf": row[COL_NF] if len(row) > COL_NF else None,
            "parcela": row[COL_PARCELA] if len(row) > COL_PARCELA else None,
            "pagto": serializa(row[COL_PAGTO]) if len(row) > COL_PAGTO else None,
            "vencto": serializa(row[COL_VENCTO]) if len(row) > COL_VENCTO else None,
            "pagos": row[COL_PAGOS] if len(row) > COL_PAGOS else None,
            "valor": row[COL_VALOR] if len(row) > COL_VALOR else None,
        }
        resultado.append(item)

    return resultado


if __name__ == "__main__":
    todos = []
    for cnpj, fname in ARQUIVOS.items():
        itens = extrair(fname, cnpj)
        print(f"{cnpj}: {len(itens)} lancamentos de despesas escritorio")
        todos.extend(itens)

    with open("fluxo-financeiro/despesas_brutas.json", "w", encoding="utf-8") as f:
        json.dump(todos, f, ensure_ascii=False, indent=2)

    print(f"TOTAL: {len(todos)} lancamentos -> fluxo-financeiro/despesas_brutas.json")
