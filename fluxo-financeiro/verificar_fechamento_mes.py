"""
Confere se o total do Fluxo Financeiro (banco) bate com o total da planilha
FLUXO FINANCEIRO - 2026.xlsx, para um mes/ano e ambiente (local ou producao).

Uso:
    python verificar_fechamento_mes.py --mes 3 --ano 2026 --ambiente local
    python verificar_fechamento_mes.py --mes 3 --ano 2026 --ambiente producao

Nao precisa abrir SSH manual nem mysql client -- o script cuida de tudo
(local: conecta direto no MySQL do Docker; producao: usa paramiko pra rodar
a query remota via SSH, mesma credencial ja usada no resto do projeto).
"""
import argparse
import calendar
import io
import sys

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")
import sys

import pymysql
import openpyxl
import paramiko

XLSX_PATH = r"C:\Users\amand\OneDrive\Documentos\CMport\cmport-system\FLUXO FINANCEIRO - 2026.xlsx"
SHEET_NAME = "Entradas e SAIDAS - 2026"
MAX_ROW = 1520  # limite conhecido da planilha (linhas com dado real; NAO usar ws.max_row, trava)

LOCAL_DB = dict(host="localhost", port=3306, user="root", password="cmport2026",
                database="cmport_gerenciamento", charset="utf8mb4")

PROD_HOST = "168.231.96.184"
PROD_SSH_KEY = None  # usa a chave default do usuario (~/.ssh/id_ed25519) via paramiko agent/config


def total_planilha(mes: int, ano: int):
    """Soma Contrato (Manutencao) e Assistencia da planilha, filtrando pela
    data de PAGAMENTO (coluna 8) caindo no mes/ano -- mesmo criterio usado na
    query do banco (MONTH(b.data_pagamento)). Regra 1 do projeto: cada linha
    da planilha representa a parcela que caiu naquele mes especifico, e o
    criterio de "cair no mes" e sempre a data em que foi pago, nao o vencimento
    (uma parcela vencida em fev mas paga em mar conta como marco).
    Usa iter_rows(min_col=1,max_col=9) -- MUITO mais rapido que ws.cell() nesse
    arquivo especifico (dimensao inflada faz .cell() individual travar/demorar)."""
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    manutencao = 0.0
    assistencia = 0.0
    n_manutencao = n_assistencia = 0
    for row in ws.iter_rows(min_row=1, max_row=MAX_ROW, min_col=1, max_col=9, values_only=True):
        categoria = row[3]
        if categoria not in ("Assistencia", "Contrato"):
            continue
        d_pag, valor = row[6], row[8]
        if not isinstance(valor, (int, float)):
            continue
        if d_pag is None or not (d_pag.month == mes and d_pag.year == ano):
            continue
        if categoria == "Contrato":
            manutencao += valor
            n_manutencao += 1
        else:
            assistencia += valor
            n_assistencia += 1
    wb.close()
    return {"manutencao": manutencao, "n_manutencao": n_manutencao,
            "assistencia": assistencia, "n_assistencia": n_assistencia}


QUERY = """
SELECT
  (SELECT COALESCE(SUM(b.valor_nominal),0) FROM notas_fiscais nf JOIN boletos b ON b.nota_fiscal_id=nf.id
     WHERE nf.tipo='MANUTENCAO' AND MONTH(b.data_pagamento)={mes} AND YEAR(b.data_pagamento)={ano}) AS manutencao,
  (SELECT COALESCE(SUM(b.valor_nominal),0) FROM notas_fiscais nf JOIN boletos b ON b.nota_fiscal_id=nf.id
     WHERE nf.tipo='ASSISTENCIA' AND MONTH(b.data_pagamento)={mes} AND YEAR(b.data_pagamento)={ano}) AS assistencia_notas,
  (SELECT COALESCE(SUM(valor),0) FROM recibos
     WHERE status='PAGO' AND MONTH(data_pagamento)={mes} AND YEAR(data_pagamento)={ano}) AS assistencia_recibos,
  (SELECT COUNT(DISTINCT nf.id) FROM notas_fiscais nf JOIN boletos b ON b.nota_fiscal_id=nf.id
     WHERE nf.tipo='ASSISTENCIA' AND MONTH(b.data_pagamento)={mes} AND YEAR(b.data_pagamento)={ano}) AS n_assistencia_notas,
  (SELECT COUNT(*) FROM recibos
     WHERE status='PAGO' AND MONTH(data_pagamento)={mes} AND YEAR(data_pagamento)={ano}) AS n_recibos,
  (SELECT COUNT(*) FROM (
     SELECT nf.id FROM notas_fiscais nf JOIN manutencoes_assistencias ma ON ma.nota_fiscal_id=nf.id
     WHERE nf.tipo='ASSISTENCIA' AND nf.id IN (
       SELECT DISTINCT b.nota_fiscal_id FROM boletos b WHERE MONTH(b.data_pagamento)={mes} AND YEAR(b.data_pagamento)={ano})
     GROUP BY nf.id HAVING COUNT(ma.id) > 1
   ) dup) AS notas_com_servico_duplicado
"""


def total_banco_local(mes: int, ano: int):
    conn = pymysql.connect(**LOCAL_DB, cursorclass=pymysql.cursors.DictCursor)
    with conn.cursor() as cur:
        cur.execute(QUERY.format(mes=mes, ano=ano))
        row = cur.fetchone()
    conn.close()
    for k in ("manutencao", "assistencia_notas", "assistencia_recibos"):
        row[k] = float(row[k])
    return row


def total_banco_producao(mes: int, ano: int):
    ssh = paramiko.SSHClient()
    ssh.set_missing_host_key_policy(paramiko.AutoAddPolicy())
    ssh.connect(PROD_HOST, username="root", timeout=15)
    query_sql = QUERY.format(mes=mes, ano=ano).replace("\n", " ")
    cmd = (
        "docker exec cmport_db sh -c "
        "'mysql -uroot -p\"$MYSQL_ROOT_PASSWORD\" -N -e \"USE $MYSQL_DATABASE; " + query_sql.replace('"', '\\"') + "\"'"
    )
    stdin, stdout, stderr = ssh.exec_command(cmd, timeout=30)
    out = stdout.read().decode("utf-8", errors="replace").strip()
    err = stderr.read().decode("utf-8", errors="replace")
    ssh.close()
    if not out:
        raise RuntimeError(f"Sem retorno de producao. stderr: {err}")
    vals = out.split("\t")
    keys = ["manutencao", "assistencia_notas", "assistencia_recibos",
            "n_assistencia_notas", "n_recibos", "notas_com_servico_duplicado"]
    return {k: (float(v) if "." in v or k in ("manutencao", "assistencia_notas", "assistencia_recibos") else int(v))
            for k, v in zip(keys, vals)}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--mes", type=int, required=True, help="1-12")
    parser.add_argument("--ano", type=int, default=2026)
    parser.add_argument("--ambiente", choices=["local", "producao"], default="local")
    args = parser.parse_args()

    nome_mes = calendar.month_name[args.mes]
    print(f"=== Fechamento de {nome_mes}/{args.ano} — ambiente: {args.ambiente} ===\n")

    print("Lendo planilha...")
    p = total_planilha(args.mes, args.ano)

    print(f"Consultando banco ({args.ambiente})...")
    b = total_banco_local(args.mes, args.ano) if args.ambiente == "local" else total_banco_producao(args.mes, args.ano)

    banco_assistencia = b["assistencia_notas"] + b["assistencia_recibos"]

    print(f"\n{'':20}{'Planilha':>15}{'Banco':>15}{'Diferenca':>15}")
    diff_manut = b["manutencao"] - p["manutencao"]
    print(f"{'Manutencao':20}{p['manutencao']:>15.2f}{b['manutencao']:>15.2f}{diff_manut:>15.2f}")
    diff_assist = banco_assistencia - p["assistencia"]
    print(f"{'Assistencia':20}{p['assistencia']:>15.2f}{banco_assistencia:>15.2f}{diff_assist:>15.2f}")
    total_p = p["manutencao"] + p["assistencia"]
    total_b = b["manutencao"] + banco_assistencia
    print(f"{'TOTAL':20}{total_p:>15.2f}{total_b:>15.2f}{total_b - total_p:>15.2f}")

    print(f"\nNotas Assistencia no banco: {b['n_assistencia_notas']} | Recibos pagos: {b['n_recibos']}")
    print(f"Linhas Assistencia na planilha: {p['n_assistencia']} | Linhas Contrato: {p['n_manutencao']}")

    if b["notas_com_servico_duplicado"] > 0:
        print(f"\n⚠️  ALERTA: {b['notas_com_servico_duplicado']} nota(s) com mais de 1 servico vinculado (bug de duplicacao)")

    if abs(diff_manut) < 0.02 and abs(diff_assist) < 0.02:
        print("\n✅ Mes fechado — banco bate exato com a planilha.")
    else:
        print(f"\n🔶 Diferenca de R$ {total_b - total_p:.2f} — revisar item a item antes de considerar fechado.")


if __name__ == "__main__":
    main()
