# -*- coding: utf-8 -*-
"""
Valida Manutencao + Assistencia + Recibos de TODOS os meses de uma vez:
planilha mestre x /financeiro/fluxo-mensal em producao, pro CNPJ escolhido.

Uso:
    ./venv/Scripts/python.exe fluxo-financeiro/validar_fluxo_todos_meses.py --empresa cmport
    ./venv/Scripts/python.exe fluxo-financeiro/validar_fluxo_todos_meses.py --empresa tec
"""
import argparse
import sys
import openpyxl
import requests

DOCS_DIR = r"C:\Users\amand\OneDrive\Documentos\CMport\cmport-system\docs-e-planilhas"
SHEET_NAME = "Entradas e SAIDAS - 2026"

# MAX_ROW checado manualmente por planilha (nao usar ws.max_row, dimensao inflada trava) --
# CMPORT: dado real ate linha 2143 (checado em 12/08/2026)
# TEC: dado real ate linha 515, comeca so em Abril/2026 (checado em 12/08/2026)
EMPRESAS = {
    "cmport": dict(
        xlsx=f"{DOCS_DIR}\\FLUXO FINANCEIRO - 2026.xlsx",
        max_row=2200,
        cnpj="22.761.557/0001-88",
    ),
    "tec": dict(
        xlsx=f"{DOCS_DIR}\\FLUXO FINANCEIRO CMPORT TEC - 2026.xlsx",
        max_row=600,
        cnpj="65756913000188",  # TEC esta cadastrado sem pontuacao no banco (CMPORT esta com)
    ),
}

BASE = "http://168.231.96.184/api/v1"
LOGIN = {"email": "atilagmoura@gmail.com", "senha": "22164855"}

ANO = 2026
MESES = range(1, 9)  # Jan-Ago; ajustar conforme a planilha crescer


def extrair_planilha(xlsx_path, max_row):
    """Retorna {mes: {'manutencao':x,'assistencia':x,'recibos':x}} a partir
    da secao Manutencoes(categoria='Contrato') + Assistencias(categoria='Assistencia'),
    somando por mes de PAGAMENTO (coluna G / indice 6 - NUNCA vencimento).
    Categoria comparada com o texto EXATO ("Contrato"/"Assistencia", com maiuscula) --
    NAO normalizar para minusculo aqui: confirmado em 12/08/2026 que tanto CMPORT
    quanto TEC usam 'assistencia' minusculo pra marcar linhas da secao Entrada/Bancos
    (transferencias internas entre contas, ex: "Recebimento de transferencia conta
    Inter CMPORT/ CMPORT TEC"), que NAO sao receita de Assistencia real. Comparar
    case-insensitive pega essas transferencias e infla o total (testado e revertido --
    quebrou meses que antes batiam exato)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[SHEET_NAME]
    dados = {m: {"manutencao": 0.0, "assistencia": 0.0, "recibos": 0.0} for m in MESES}
    for row in ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=9, values_only=True):
        categoria, nf, d_pag, valor = row[3], row[4], row[6], row[8]
        if categoria not in ("Contrato", "Assistencia"):
            continue
        if not isinstance(valor, (int, float)):
            continue
        if d_pag is None or d_pag.year != ANO or d_pag.month not in MESES:
            continue
        if d_pag.year < 2000:  # datas corrompidas tipo 1900/1901 (bug de parse do Excel)
            continue
        mes = d_pag.month
        nf_norm = str(nf).strip().lower() if nf else None
        if nf_norm in ("pix", "recibo") or nf is None:
            dados[mes]["recibos"] += valor
        elif categoria == "Contrato":
            dados[mes]["manutencao"] += valor
        else:
            dados[mes]["assistencia"] += valor
    wb.close()
    return {m: {k: round(v, 2) for k, v in d.items()} for m, d in dados.items()}


def extrair_sistema(cnpj):
    token = requests.post(f"{BASE}/auth/login", json=LOGIN).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}
    dados = {}
    for mes in MESES:
        r = requests.get(f"{BASE}/financeiro/fluxo-mensal", params={"ano": ANO, "mes": mes}, headers=headers)
        r.raise_for_status()
        resp = r.json()
        empresa = next((c for c in resp["cnpjs"] if c["cnpj"].strip() == cnpj), None)
        if empresa:
            dados[mes] = {
                "manutencao": empresa["total_manutencao"],
                # planilha so tem categoria "Assistencia" (sem distincao Produto/Servico) --
                # soma total_produto aqui pra comparar contra o mesmo agregado (ver commit
                # b1a67d3, fluxo_financeiro_service.py passou a contar notas tipo PRODUTO)
                "assistencia": round(empresa["total_assistencia"] + empresa.get("total_produto", 0.0), 2),
                "recibos": empresa["total_recibos"],
            }
        else:
            dados[mes] = {"manutencao": 0.0, "assistencia": 0.0, "recibos": 0.0}
    return dados


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--empresa", choices=list(EMPRESAS), default="cmport")
    args = parser.parse_args()
    cfg = EMPRESAS[args.empresa]

    print(f"=== Empresa: {args.empresa.upper()} | planilha: {cfg['xlsx']} ===\n")

    planilha = extrair_planilha(cfg["xlsx"], cfg["max_row"])
    sistema = extrair_sistema(cfg["cnpj"])

    nomes_mes = ["", "Jan", "Fev", "Mar", "Abr", "Mai", "Jun", "Jul", "Ago", "Set", "Out", "Nov", "Dez"]
    total_diff_geral = 0.0
    linhas_com_diff = []

    print(f"{'Mes':5} {'Categoria':11} {'Planilha':>12} {'Sistema':>12} {'Diferenca':>12}")
    print("-" * 56)
    for mes in MESES:
        p, s = planilha[mes], sistema[mes]
        for cat in ("manutencao", "assistencia", "recibos"):
            diff = round(s[cat] - p[cat], 2)
            marca = "" if abs(diff) < 0.01 else "  <<<"
            print(f"{nomes_mes[mes]:5} {cat:11} {p[cat]:>12.2f} {s[cat]:>12.2f} {diff:>12.2f}{marca}")
            if abs(diff) >= 0.01:
                linhas_com_diff.append((nomes_mes[mes], cat, p[cat], s[cat], diff))
                total_diff_geral += diff
        tot_p = round(sum(p.values()), 2)
        tot_s = round(sum(s.values()), 2)
        diff_tot = round(tot_s - tot_p, 2)
        marca = "" if abs(diff_tot) < 0.01 else "  <<<"
        print(f"{nomes_mes[mes]:5} {'TOTAL':11} {tot_p:>12.2f} {tot_s:>12.2f} {diff_tot:>12.2f}{marca}")
        print("-" * 56)

    print(f"\n{len(linhas_com_diff)} linha(s) com diferenca >= R$0,01:")
    for nm, cat, p, s, diff in linhas_com_diff:
        print(f"  {nm} / {cat}: planilha R$ {p:.2f} vs sistema R$ {s:.2f} (diff R$ {diff:.2f})")


if __name__ == "__main__":
    sys.exit(main())
